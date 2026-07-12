# -*- coding: utf-8 -*-
"""
Permanent transaction ledger for all credit balance changes.
Every balance mutation (recharge code, hint purchase, admin action)
is recorded here and never deleted automatically.
"""
import io
import json
import logging
import os
import uuid
from datetime import datetime, timezone

from config import TRANSACTIONS_FILE

logger = logging.getLogger(__name__)

TYPE_LABELS: dict[str, str] = {
    "recharge_code": "شحن بكود",
    "hint_purchase":  "كشف حرف",
    "word_reward":    "مكافأة فتح كلمة",
    "admin_add":      "إضافة بواسطة المشرف",
    "admin_remove":   "خصم بواسطة المشرف",
    "admin_reset":    "تصفير بواسطة المشرف",
    "quiz_result_add":    "نتيجة اختبار",
    "quiz_result_remove": "إزالة نتيجة اختبار سابقة",
}

_ARROW: dict[str, str] = {
    "recharge_code": "➕",
    "hint_purchase":  "➖",
    "word_reward":    "➕",
    "admin_add":      "➕",
    "admin_remove":   "➖",
    "admin_reset":    "🔄",
    "quiz_result_add":    "🏆",
    "quiz_result_remove": "♻️",
}

_SIGN: dict[str, str] = {
    "recharge_code": "+",
    "hint_purchase":  "-",
    "word_reward":    "+",
    "admin_add":      "+",
    "admin_remove":   "-",
    "admin_reset":    "±",
    "quiz_result_add":    "+",
    "quiz_result_remove": "-",
}


# ── Storage ────────────────────────────────────────────────────────────────────

def _load() -> list:
    os.makedirs(os.path.dirname(TRANSACTIONS_FILE), exist_ok=True)
    if not os.path.exists(TRANSACTIONS_FILE):
        return []
    try:
        with open(TRANSACTIONS_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except (json.JSONDecodeError, OSError):
        return []


def _save(data: list) -> None:
    os.makedirs(os.path.dirname(TRANSACTIONS_FILE), exist_ok=True)
    with open(TRANSACTIONS_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, indent=2, ensure_ascii=False)


# ── Public API ─────────────────────────────────────────────────────────────────

def record(user_id: int, full_name: str, txn_type: str,
           amount: int, balance_before: int, balance_after: int,
           description: str = "") -> dict:
    """Append one transaction to the ledger and return the entry."""
    now   = datetime.now(timezone.utc)
    entry = {
        "txn_id":         str(uuid.uuid4()),
        "timestamp":      now.isoformat(),
        "date":           now.strftime("%d/%m/%Y"),
        "time":           now.strftime("%H:%M"),
        "user_id":        str(user_id),
        "full_name":      full_name or "—",
        "type":           txn_type,
        "type_label":     TYPE_LABELS.get(txn_type, txn_type),
        "amount":         amount,
        "balance_before": balance_before,
        "balance_after":  balance_after,
        "description":    description or TYPE_LABELS.get(txn_type, txn_type),
    }
    txns = _load()
    txns.append(entry)
    _save(txns)
    return entry


def load_all() -> list:
    return _load()


def load_user_txns(user_id: int) -> list:
    uid = str(user_id)
    return [t for t in _load() if t.get("user_id") == uid]


def delete_all() -> None:
    _save([])


def delete_user_txns(user_id: int) -> int:
    """Permanently remove all transactions belonging to one user."""
    uid  = str(user_id)
    txns = _load()
    kept = [t for t in txns if t.get("user_id") != uid]
    removed = len(txns) - len(kept)
    if removed:
        _save(kept)
    return removed


def rename_user_txns(user_id: int, new_name: str) -> int:
    """Update full_name in every transaction belonging to one user. Returns count updated."""
    uid  = str(user_id)
    txns = _load()
    count = 0
    for t in txns:
        if t.get("user_id") == uid:
            t["full_name"] = new_name
            count += 1
    if count:
        _save(txns)
    return count


def search_users_in_txns(query: str) -> list[str]:
    """Return unique user_ids whose name or UID matches query."""
    q    = query.strip().lower()
    seen: set[str] = set()
    out:  list[str] = []
    for t in _load():
        uid  = t.get("user_id", "")
        name = (t.get("full_name") or "").lower()
        if uid not in seen and (q in name or uid == q):
            seen.add(uid)
            out.append(uid)
    return out


# ── Display ────────────────────────────────────────────────────────────────────

def format_entry(t: dict) -> str:
    arrow  = _ARROW.get(t.get("type", ""), "💳")
    sign   = _SIGN.get(t.get("type", ""), "")
    amount = t.get("amount", 0)
    desc   = t.get("description") or t.get("type_label", "—")
    bef    = t.get("balance_before", 0)
    aft    = t.get("balance_after",  0)
    date   = t.get("date", "—")
    time_  = t.get("time", "")
    return (
        f"📅 {date}  🕐 {time_}\n"
        f"{arrow} {sign}{amount} نقطة\n"
        f"{desc}\n\n"
        f"الرصيد:\n{bef} ← {aft}"
    )


# ── Export: Excel ──────────────────────────────────────────────────────────────

def export_excel(transactions: list) -> io.BytesIO:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "سجل الحركات"
    ws.sheet_view.rightToLeft = True

    headers = [
        "#", "التاريخ", "الوقت", "معرف المستخدم", "الاسم الكامل",
        "نوع العملية", "المبلغ", "الرصيد قبل", "الرصيد بعد", "الوصف",
    ]
    h_fill = PatternFill("solid", fgColor="2E4057")
    h_font = Font(bold=True, color="FFFFFF")
    ws.append(headers)
    for cell in ws[1]:
        cell.fill      = h_fill
        cell.font      = h_font
        cell.alignment = Alignment(horizontal="center")

    for i, t in enumerate(transactions, 1):
        ws.append([
            i,
            t.get("date", ""),
            t.get("time", ""),
            t.get("user_id", ""),
            t.get("full_name", ""),
            t.get("type_label", t.get("type", "")),
            t.get("amount", 0),
            t.get("balance_before", 0),
            t.get("balance_after",  0),
            t.get("description", ""),
        ])

    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 45)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ── Export: PDF ────────────────────────────────────────────────────────────────

_FONT_PATH = "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf"


def export_pdf(transactions: list) -> io.BytesIO:
    from fpdf import FPDF

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=12)
    pdf.add_page()

    if os.path.exists(_FONT_PATH):
        pdf.add_font("Uni", style="", fname=_FONT_PATH)
        fn = "Uni"
    else:
        fn = "Helvetica"

    pdf.set_font(fn, size=14)
    pdf.set_fill_color(46, 64, 87)
    pdf.set_text_color(255, 255, 255)
    pdf.cell(0, 10, "Credit Transaction History - سجل حركة الرصيد",
             border=0, new_x="LMARGIN", new_y="NEXT", align="C", fill=True)
    pdf.ln(3)

    cols   = ["#", "Date", "Time", "User ID", "Name", "Type", "Amt", "Before", "After", "Description"]
    widths = [10, 22, 14, 26, 35, 42, 14, 16, 16, 72]

    pdf.set_font(fn, size=8)
    pdf.set_fill_color(180, 210, 230)
    pdf.set_text_color(0, 0, 0)
    for col, w in zip(cols, widths):
        pdf.cell(w, 8, col, border=1, align="C", fill=True)
    pdf.ln()

    pdf.set_font(fn, size=7)
    for i, t in enumerate(transactions, 1):
        alt = (i % 2 == 0)
        pdf.set_fill_color(245, 248, 252) if alt else pdf.set_fill_color(255, 255, 255)
        row = [
            str(i),
            t.get("date", ""),
            t.get("time", ""),
            t.get("user_id", ""),
            (t.get("full_name") or "—")[:22],
            (t.get("type_label") or t.get("type", ""))[:22],
            str(t.get("amount", 0)),
            str(t.get("balance_before", 0)),
            str(t.get("balance_after",  0)),
            (t.get("description") or "")[:38],
        ]
        for val, w in zip(row, widths):
            pdf.cell(w, 6, str(val), border=1, fill=True)
        pdf.ln()

    return io.BytesIO(pdf.output())
