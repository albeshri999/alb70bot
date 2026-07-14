# -*- coding: utf-8 -*-
"""
Admin side of '👥 اختبار توزيع الفرق' (Team-Distribution Test) — supports the
exact same feature set as the regular '📝 إدارة الاختبارات' system (creation
wizard, full field/question editing, timed tests, visibility), PLUS its own
extra features: an independent "entry open/closed" gate, "👥 تقسيم الفرق"
(balanced team splitting via snake draft), and Excel export of the split.

Design (same proven pattern as quiz_admin.py / admin_settings.py):
- Its own, fully independent ConversationHandler, entered via the
  "adm_distro" callback button (the only line added to admin.py's keyboard —
  see admin.py's _main_kb()).
- Because python-telegram-bot's ConversationHandler ignores updates that
  don't match its current state instead of consuming them, pressing
  "⬅️ القائمة الرئيسية" (callback_data="adm_main") from anywhere in this
  conversation is NOT handled here — it falls through untouched to the
  original admin ConversationHandler (still parked in its MAIN state),
  which shows the normal admin main menu exactly as before.
- All data lives in distro_storage.py (its own JSON files). Nothing here
  reads or writes days.json, users.json, config.json, credit_log.json,
  transactions.json, recharge_codes.json, admin_log.json, quizzes.json,
  quiz_results.json, quiz_sessions.json, or quiz_credit_log.json.
- No points are ever added to any participant's balance from this feature —
  there is no call anywhere in this file to credits.add_credits() or
  transactions.record(). Scores (now weighted by "درجة السؤال", same as the
  regular quiz system) exist ONLY to rank participants for the team split;
  this test never appears on the leaderboard.
"""
import io
import logging

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ContextTypes, ConversationHandler, CommandHandler,
    MessageHandler, CallbackQueryHandler, filters,
)
from telegram.constants import ParseMode

import admins_store
import distro_storage as ds

logger = logging.getLogger(__name__)

# ── States ────────────────────────────────────────────────────────────────────
(DT_HUB, DT_LIST, DT_QUIZ_MENU, DT_DEL_CONFIRM,
 DT_C_NAME, DT_C_DESC, DT_C_POINTS, DT_C_TIMED, DT_C_MINUTES, DT_C_VISIBLE,
 DT_C_Q_TEXT, DT_C_Q_OPT_A, DT_C_Q_OPT_B, DT_C_Q_CORRECT, DT_C_Q_MORE,
 DT_EDIT_MENU, DT_E_NAME, DT_E_DESC, DT_E_POINTS, DT_E_TIMED, DT_E_MINUTES,
 DT_E_Q_LIST, DT_E_Q_FIELD_MENU, DT_E_Q_FIELD_VAL, DT_E_Q_DEL_CONFIRM,
 DT_SPLIT_SIZE, DT_SPLIT_VIEW,
 ) = range(27)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _is_admin(uid: int) -> bool:
    return admins_store.is_admin(uid)


def _md(text: str) -> str:
    for ch in ("_", "*", "`", "[", "]"):
        text = str(text).replace(ch, f"\\{ch}")
    return text


async def _reply(update: Update, text: str, keyboard=None):
    kw = {"text": text, "parse_mode": ParseMode.MARKDOWN}
    if keyboard:
        kw["reply_markup"] = keyboard
    if update.callback_query:
        try:
            await update.callback_query.message.edit_text(**kw)
        except Exception as e1:
            logger.warning("_reply edit_text failed: %s", e1)
            try:
                await update.callback_query.message.reply_text(**kw)
            except Exception as e2:
                logger.error("_reply reply_text also failed: %s", e2)
    else:
        await update.message.reply_text(**kw)


def _yn(yes_cb: str, no_cb: str) -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ نعم", callback_data=yes_cb),
        InlineKeyboardButton("❌ لا", callback_data=no_cb),
    ]])


# ── Hub ───────────────────────────────────────────────────────────────────────

def _hub_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ إنشاء اختبار توزيع", callback_data="dt_create")],
        [InlineKeyboardButton("📋 قائمة اختبارات التوزيع", callback_data="dt_list_menu")],
        [InlineKeyboardButton("👥 تقسيم الفرق", callback_data="dt_list_split")],
        [InlineKeyboardButton("📊 نتائج اختبار التوزيع", callback_data="dt_list_results")],
        [InlineKeyboardButton("🗑 حذف اختبار", callback_data="dt_list_delete")],
        [InlineKeyboardButton("⬅️ القائمة الرئيسية", callback_data="adm_main")],
    ])


async def dt_hub(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not _is_admin(update.effective_user.id):
        if update.callback_query:
            await update.callback_query.answer("⛔ ليس لديك صلاحية.", show_alert=True)
        return ConversationHandler.END
    if update.callback_query:
        await update.callback_query.answer()
    context.user_data.pop("dt_quiz_id", None)
    context.user_data.pop("dt_action", None)
    context.user_data.pop("dt_new", None)
    context.user_data.pop("dt_new_q", None)
    context.user_data.pop("dt_mode", None)
    await _reply(update, "👥 *اختبار توزيع الفرق*\n\nاختر العملية:", _hub_kb())
    return DT_HUB


# ── Quiz list (shared by menu/delete/results/split actions) ─────────────────

def _quiz_list_kb() -> InlineKeyboardMarkup:
    quizzes = ds.load_quizzes()
    rows = [
        [InlineKeyboardButton(q.get("name", "—"), callback_data=f"dt_pick_{qid}")]
        for qid, q in sorted(quizzes.items(), key=lambda kv: int(kv[0]))
    ]
    rows.append([InlineKeyboardButton("🔙 رجوع", callback_data="dt_hub")])
    return InlineKeyboardMarkup(rows)


async def dt_list_entry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query  = update.callback_query
    await query.answer()
    action = query.data.replace("dt_list_", "")  # menu | delete | results | split
    context.user_data["dt_action"] = action

    quizzes = ds.load_quizzes()
    if not quizzes:
        await _reply(
            update, "📭 لا توجد اختبارات توزيع بعد.\n\nأنشئ اختباراً جديداً أولاً.",
            InlineKeyboardMarkup([[InlineKeyboardButton("🔙 رجوع", callback_data="dt_hub")]]),
        )
        return DT_LIST

    await _reply(update, "📋 *اختبارات توزيع الفرق*\n\nاختر اختباراً:", _quiz_list_kb())
    return DT_LIST


async def dt_pick(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    await query.answer()
    quiz_id = query.data.replace("dt_pick_", "")
    context.user_data["dt_quiz_id"] = quiz_id
    action = context.user_data.get("dt_action", "menu")

    if action == "delete":
        return await _show_delete_confirm(update, context)
    if action == "results":
        return await _show_results(update, context)
    if action == "split":
        return await _ask_team_size(update, context)
    return await _show_quiz_menu(update, context)


# ── Per-quiz menu ─────────────────────────────────────────────────────────────

def _quiz_menu_kb(quiz: dict) -> InlineKeyboardMarkup:
    vis   = "👁 إخفاء الاختبار" if quiz.get("visible") else "👁 إظهار الاختبار"
    entry = "🔒 منع الدخول" if ds.is_entry_open(quiz) else "🔓 السماح بالدخول"
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("✏️ تعديل", callback_data="dt_edit_menu")],
        [InlineKeyboardButton(vis, callback_data="dt_toggle_visible")],
        [InlineKeyboardButton(entry, callback_data="dt_toggle_entry")],
        [InlineKeyboardButton("📊 النتائج", callback_data="dt_show_results")],
        [InlineKeyboardButton("👥 تقسيم الفرق", callback_data="dt_goto_split")],
        [InlineKeyboardButton("🗑 حذف", callback_data="dt_delete_confirm")],
        [InlineKeyboardButton("🔙 رجوع لقائمة اختبارات التوزيع", callback_data="dt_list_menu")],
        [InlineKeyboardButton("⬅️ القائمة الرئيسية", callback_data="adm_main")],
    ])


def _quiz_summary(quiz: dict) -> str:
    n_q   = len(quiz.get("questions", []))
    total = n_q * int(quiz.get("points_per_question", 0))
    time_line = f"{quiz.get('time_minutes')} دقيقة" if quiz.get("timed") else "بدون تحديد"
    n_results = len(ds.results_for_quiz(quiz.get("id")))
    lines = [f"👥 *{_md(quiz.get('name', '—'))}*"]
    if quiz.get("description"):
        lines.append(f"_{_md(quiz['description'])}_")
    lines += [
        "",
        f"❓ عدد الأسئلة: *{n_q}*",
        f"🔢 درجة كل سؤال: *{quiz.get('points_per_question', 0)}*  (الإجمالي: *{total}*)",
        f"⏱ الوقت: *{time_line}*",
        f"👁 ظاهر للمتسابقين: *{'نعم' if quiz.get('visible') else 'لا'}*",
        f"🚪 الدخول: *{'مسموح' if ds.is_entry_open(quiz) else 'ممنوع'}*",
        f"✅ عدد من أنهى الاختبار: *{n_results}*",
    ]
    return "\n".join(lines)


async def _show_quiz_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    quiz_id = context.user_data.get("dt_quiz_id")
    quiz    = ds.get_quiz(quiz_id)
    if not quiz:
        await _reply(update, "⚠️ لم يعد هذا الاختبار موجوداً.", _hub_kb())
        return DT_HUB
    await _reply(update, _quiz_summary(quiz) + "\n\nاختر العملية:", _quiz_menu_kb(quiz))
    return DT_QUIZ_MENU


async def dt_back_to_quiz_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    return await _show_quiz_menu(update, context)


async def dt_toggle_visible(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    quiz_id = context.user_data.get("dt_quiz_id")
    quiz    = ds.get_quiz(quiz_id)
    ds.set_quiz_visible(quiz_id, not quiz.get("visible"))
    return await _show_quiz_menu(update, context)


async def dt_toggle_entry(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    quiz_id = context.user_data.get("dt_quiz_id")
    quiz    = ds.get_quiz(quiz_id)
    ds.set_entry_open(quiz_id, not ds.is_entry_open(quiz))
    return await _show_quiz_menu(update, context)


async def dt_goto_split(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    return await _ask_team_size(update, context)


# ── Delete quiz ───────────────────────────────────────────────────────────────

async def _show_delete_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    quiz_id = context.user_data.get("dt_quiz_id")
    quiz    = ds.get_quiz(quiz_id)
    if not quiz:
        await _reply(update, "⚠️ لم يعد هذا الاختبار موجوداً.", _hub_kb())
        return DT_HUB
    await _reply(
        update,
        f"🗑 هل أنت متأكد من حذف اختبار التوزيع:\n*{_md(quiz.get('name'))}*؟\n\n"
        "سيتم حذف جميع أسئلته ونتائجه وتقسيماته المحفوظة نهائياً.",
        _yn("dt_delete_yes", "dt_back_to_quiz_menu"),
    )
    return DT_DEL_CONFIRM


async def dt_delete_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    return await _show_delete_confirm(update, context)


async def dt_delete_yes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    quiz_id = context.user_data.get("dt_quiz_id")
    ds.delete_quiz(quiz_id)
    context.user_data.pop("dt_quiz_id", None)
    await _reply(update, "✅ تم حذف اختبار التوزيع بنجاح.", _hub_kb())
    return DT_HUB


# ── Results ───────────────────────────────────────────────────────────────────

def _results_text(quiz_id) -> str:
    quiz    = ds.get_quiz(quiz_id)
    results = sorted(ds.results_for_quiz(quiz_id), key=lambda r: -int(r.get("score", 0)))
    lines = [f"📊 *نتائج اختبار التوزيع: {_md(quiz.get('name', '—'))}*\n"]
    if not results:
        lines.append("لا توجد نتائج بعد.")
        return "\n".join(lines)
    for i, r in enumerate(results, 1):
        lines.append(
            f"{i}. *{_md(r.get('user_name', '—'))}*  —  "
            f"الدرجة: {r.get('score', 0)}/{r.get('total_score', 0)}"
        )
    return "\n".join(lines)


async def _show_results(update: Update, context: ContextTypes.DEFAULT_TYPE):
    quiz_id = context.user_data.get("dt_quiz_id")
    if not ds.get_quiz(quiz_id):
        await _reply(update, "⚠️ لم يعد هذا الاختبار موجوداً.", _hub_kb())
        return DT_HUB
    kb = InlineKeyboardMarkup([[InlineKeyboardButton("🔙 رجوع", callback_data="dt_back_to_quiz_menu")]])
    await _reply(update, _results_text(quiz_id), kb)
    return DT_QUIZ_MENU


async def dt_show_results(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    return await _show_results(update, context)


# ── Team split ("👥 تقسيم الفرق") ─────────────────────────────────────────────

async def _ask_team_size(update: Update, context: ContextTypes.DEFAULT_TYPE):
    quiz_id = context.user_data.get("dt_quiz_id")
    quiz    = ds.get_quiz(quiz_id)
    if not quiz:
        await _reply(update, "⚠️ لم يعد هذا الاختبار موجوداً.", _hub_kb())
        return DT_HUB
    if not ds.results_for_quiz(quiz_id):
        await _reply(
            update,
            f"📭 لا يوجد أي متسابق أنهى اختبار *{_md(quiz.get('name'))}* بعد.\n\n"
            "انتظر حتى ينتهي المتسابقون من الاختبار قبل التقسيم.",
            InlineKeyboardMarkup([[InlineKeyboardButton("🔙 رجوع", callback_data="dt_hub")]]),
        )
        return DT_HUB
    await _reply(update, "🔢 كم عدد أعضاء كل فريق؟\n\nمثال: 5")
    return DT_SPLIT_SIZE


def _split_text(quiz_name: str, teams: list) -> str:
    lines = [f"👥 *تقسيم فرق: {_md(quiz_name)}*\n"]
    for team in teams:
        lines.append(f"\n🔹 *{team['name']}*")
        for m in team["members"]:
            lines.append(f"   • {_md(m['user_name'])} — {m['score']} نقطة")
        lines.append(f"   ➖ المجموع: *{team['total']}*  |  المتوسط: *{team['average']}*")
    return "\n".join(lines)


def _split_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📄 تصدير الفرق", callback_data="dt_export_teams")],
        [InlineKeyboardButton("🔁 إعادة التقسيم", callback_data="dt_resplit")],
        [InlineKeyboardButton("🔙 رجوع", callback_data="dt_hub")],
        [InlineKeyboardButton("⬅️ القائمة الرئيسية", callback_data="adm_main")],
    ])


async def dt_split_size_val(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text.isdigit() or int(text) <= 0:
        await update.message.reply_text("⚠️ الرجاء إدخال رقم صحيح أكبر من صفر.")
        return DT_SPLIT_SIZE

    quiz_id = context.user_data.get("dt_quiz_id")
    quiz    = ds.get_quiz(quiz_id)
    if not quiz:
        await update.message.reply_text("⚠️ لم يعد هذا الاختبار موجوداً.", reply_markup=_hub_kb())
        return DT_HUB

    teams = ds.compute_teams(quiz_id, int(text))
    if not teams:
        await update.message.reply_text(
            "📭 لا توجد نتائج كافية لتقسيم الفرق بعد.", reply_markup=_hub_kb(),
        )
        return DT_HUB

    ds.set_team_split(quiz_id, int(text), teams)
    await update.message.reply_text(
        _split_text(quiz.get("name", "—"), teams),
        reply_markup=_split_kb(),
        parse_mode=ParseMode.MARKDOWN,
    )
    return DT_SPLIT_VIEW


async def dt_resplit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(update, "🔢 كم عدد أعضاء كل فريق؟\n\nمثال: 5")
    return DT_SPLIT_SIZE


async def dt_export_teams(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query   = update.callback_query
    await query.answer("جاري إنشاء الملف…")
    quiz_id = context.user_data.get("dt_quiz_id")
    quiz    = ds.get_quiz(quiz_id)
    split   = ds.get_team_split(quiz_id)
    if not quiz or not split:
        await query.message.reply_text("⚠️ لا يوجد تقسيم محفوظ لتصديره. استخدم 🔁 إعادة التقسيم أولاً.")
        return DT_SPLIT_VIEW

    # Lazy import: keeps this feature's failure isolated to the export button
    # (rather than crashing the whole bot at startup) if openpyxl is missing.
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    ws = wb.active
    ws.title = "تقسيم الفرق"
    ws.sheet_view.rightToLeft = True

    headers = ["اسم الفريق", "اسم المتسابق", "درجة اختبار التوزيع", "مجموع الفريق"]
    h_fill = PatternFill("solid", fgColor="1F4E79")
    h_font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    for team in split["teams"]:
        for m in team["members"]:
            ws.append([team["name"], m["user_name"], m["score"], team["total"]])

    for idx, w in enumerate([20, 28, 22, 16], 1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await query.message.reply_document(
        document=buf,
        filename=f"teams_{quiz.get('name','distro')}.xlsx",
        caption="📄 تم تصدير تقسيم الفرق.",
    )
    return DT_SPLIT_VIEW


# ── Create new distribution test (mirrors the regular quiz wizard exactly) ──

async def dt_create(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    context.user_data["dt_new"] = {}
    await _reply(update, "📝 أرسل *اسم اختبار التوزيع*:\n\nمثال: اختبار توزيع فرق المخيم")
    return DT_C_NAME


async def dt_c_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if len(text) < 2:
        await update.message.reply_text("⚠️ الرجاء إدخال اسم صحيح.")
        return DT_C_NAME
    context.user_data["dt_new"]["name"] = text
    await update.message.reply_text("🗒 أرسل *وصف الاختبار* (اختياري — أرسل `-` للتخطي):",
                                     parse_mode=ParseMode.MARKDOWN)
    return DT_C_DESC


async def dt_c_desc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    context.user_data["dt_new"]["description"] = "" if text == "-" else text
    await update.message.reply_text("🔢 أرسل *درجة كل سؤال* (رقم):\n\nمثال: 5")
    return DT_C_POINTS


async def dt_c_points(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text.isdigit() or int(text) <= 0:
        await update.message.reply_text("⚠️ الرجاء إدخال رقم صحيح أكبر من صفر.")
        return DT_C_POINTS
    context.user_data["dt_new"]["points_per_question"] = int(text)
    await update.message.reply_text("⏱ هل الاختبار محدد بوقت؟", reply_markup=_yn("dt_timed_yes", "dt_timed_no"))
    return DT_C_TIMED


async def dt_c_timed_no(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    context.user_data["dt_new"]["timed"] = False
    context.user_data["dt_new"]["time_minutes"] = None
    await _reply(update, "👁 هل يظهر الاختبار للمتسابقين الآن؟", _yn("dt_vis_yes", "dt_vis_no"))
    return DT_C_VISIBLE


async def dt_c_timed_yes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    context.user_data["dt_new"]["timed"] = True
    await _reply(update, "⏱ كم دقيقة؟\n\nمثال: 30")
    return DT_C_MINUTES


async def dt_c_minutes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text.isdigit() or int(text) <= 0:
        await update.message.reply_text("⚠️ الرجاء إدخال رقم دقائق صحيح.")
        return DT_C_MINUTES
    context.user_data["dt_new"]["time_minutes"] = int(text)
    await update.message.reply_text("👁 هل يظهر الاختبار للمتسابقين الآن؟", reply_markup=_yn("dt_vis_yes", "dt_vis_no"))
    return DT_C_VISIBLE


async def _finish_create_quiz(update: Update, context: ContextTypes.DEFAULT_TYPE, visible: bool):
    data = context.user_data.get("dt_new", {})
    quiz_id = ds.create_quiz(
        name=data.get("name", "بدون اسم"),
        description=data.get("description", ""),
        points_per_question=data.get("points_per_question", 1),
        timed=data.get("timed", False),
        time_minutes=data.get("time_minutes"),
        visible=visible,
    )
    context.user_data["dt_quiz_id"] = quiz_id
    context.user_data["dt_mode"] = "create"
    context.user_data.pop("dt_new", None)
    await _reply(
        update,
        "✅ تم إنشاء اختبار التوزيع. الآن أضف الأسئلة.\n\n"
        "📝 أرسل *نص السؤال الأول*:",
    )
    return DT_C_Q_TEXT


async def dt_c_visible_yes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    return await _finish_create_quiz(update, context, True)


async def dt_c_visible_no(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    return await _finish_create_quiz(update, context, False)


# ── Add-question loop (shared between "create new test" and "edit → add question") ──

async def dt_q_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text:
        await update.message.reply_text("⚠️ الرجاء إدخال نص السؤال.")
        return DT_C_Q_TEXT
    context.user_data["dt_new_q"] = {"question": text}
    await update.message.reply_text("أ) أرسل *نص الخيار الأول*:")
    return DT_C_Q_OPT_A


async def dt_q_opt_a(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    context.user_data["dt_new_q"]["option_a"] = text
    await update.message.reply_text("ب) أرسل *نص الخيار الثاني*:")
    return DT_C_Q_OPT_B


async def dt_q_opt_b(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    context.user_data["dt_new_q"]["option_b"] = text
    q = context.user_data["dt_new_q"]
    await update.message.reply_text(
        f"اختر *الإجابة الصحيحة*:\n\nأ) {q['option_a']}\nب) {q['option_b']}",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("✅ أ", callback_data="dt_correct_a"),
            InlineKeyboardButton("✅ ب", callback_data="dt_correct_b"),
        ]]),
    )
    return DT_C_Q_CORRECT


async def _save_new_question(update: Update, context: ContextTypes.DEFAULT_TYPE, correct: str):
    quiz_id = context.user_data.get("dt_quiz_id")
    q = context.user_data.get("dt_new_q", {})
    ds.add_question(quiz_id, q.get("question", ""), q.get("option_a", ""), q.get("option_b", ""), correct)
    context.user_data.pop("dt_new_q", None)
    await _reply(
        update,
        "✅ تمت إضافة السؤال.\n\nهل تريد إضافة سؤال آخر؟",
        InlineKeyboardMarkup([[
            InlineKeyboardButton("➕ سؤال آخر", callback_data="dt_q_more_yes"),
            InlineKeyboardButton("✅ إنهاء الاختبار", callback_data="dt_q_more_done"),
        ]]),
    )
    return DT_C_Q_MORE


async def dt_correct_a(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    return await _save_new_question(update, context, "a")


async def dt_correct_b(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    return await _save_new_question(update, context, "b")


async def dt_q_more_yes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(update, "📝 أرسل *نص السؤال* التالي:")
    return DT_C_Q_TEXT


async def dt_q_more_done(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    if context.user_data.get("dt_mode") == "edit_add":
        return await _show_question_list(update, context)
    return await _show_quiz_menu(update, context)


# ── Edit existing test ────────────────────────────────────────────────────────

def _edit_menu_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📝 اسم الاختبار", callback_data="dt_e_name")],
        [InlineKeyboardButton("🗒 الوصف", callback_data="dt_e_desc")],
        [InlineKeyboardButton("🔢 درجة السؤال", callback_data="dt_e_points")],
        [InlineKeyboardButton("⏱ الوقت", callback_data="dt_e_timed")],
        [InlineKeyboardButton("❓ الأسئلة", callback_data="dt_e_questions")],
        [InlineKeyboardButton("🔙 رجوع", callback_data="dt_back_to_quiz_menu")],
    ])


async def dt_edit_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(update, "✏️ *تعديل اختبار التوزيع*\n\nاختر ما تريد تعديله:", _edit_menu_kb())
    return DT_EDIT_MENU


async def dt_e_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(update, "📝 أرسل الاسم الجديد للاختبار:")
    return DT_E_NAME


async def dt_e_name_val(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if len(text) < 2:
        await update.message.reply_text("⚠️ الرجاء إدخال اسم صحيح.")
        return DT_E_NAME
    ds.update_quiz_field(context.user_data.get("dt_quiz_id"), name=text)
    await update.message.reply_text("✅ تم تحديث اسم الاختبار.")
    return await dt_edit_menu_direct(update, context)


async def dt_e_desc(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(update, "🗒 أرسل الوصف الجديد (أرسل `-` لإفراغه):")
    return DT_E_DESC


async def dt_e_desc_val(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    ds.update_quiz_field(context.user_data.get("dt_quiz_id"), description="" if text == "-" else text)
    await update.message.reply_text("✅ تم تحديث الوصف.")
    return await dt_edit_menu_direct(update, context)


async def dt_e_points(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(update, "🔢 أرسل درجة كل سؤال الجديدة (رقم):")
    return DT_E_POINTS


async def dt_e_points_val(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text.isdigit() or int(text) <= 0:
        await update.message.reply_text("⚠️ الرجاء إدخال رقم صحيح أكبر من صفر.")
        return DT_E_POINTS
    ds.update_quiz_field(context.user_data.get("dt_quiz_id"), points_per_question=int(text))
    await update.message.reply_text("✅ تم تحديث درجة السؤال.")
    return await dt_edit_menu_direct(update, context)


async def dt_e_timed(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(update, "⏱ هل الاختبار محدد بوقت؟", _yn("dt_e_timed_yes", "dt_e_timed_no"))
    return DT_E_TIMED


async def dt_e_timed_no(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    ds.update_quiz_field(context.user_data.get("dt_quiz_id"), timed=False, time_minutes=None)
    await _reply(update, "✅ تم إلغاء تحديد الوقت.")
    return await dt_edit_menu_direct(update, context)


async def dt_e_timed_yes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(update, "⏱ كم دقيقة؟")
    return DT_E_MINUTES


async def dt_e_minutes_val(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()
    if not text.isdigit() or int(text) <= 0:
        await update.message.reply_text("⚠️ الرجاء إدخال رقم دقائق صحيح.")
        return DT_E_MINUTES
    ds.update_quiz_field(context.user_data.get("dt_quiz_id"), timed=True, time_minutes=int(text))
    await update.message.reply_text("✅ تم تحديث وقت الاختبار.")
    return await dt_edit_menu_direct(update, context)


async def dt_edit_menu_direct(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Re-show the edit menu after a text-based field update (no callback_query)."""
    await update.message.reply_text("✏️ *تعديل اختبار التوزيع*\n\nاختر ما تريد تعديله:",
                                     reply_markup=_edit_menu_kb(), parse_mode=ParseMode.MARKDOWN)
    return DT_EDIT_MENU


# ── Questions management inside edit ──────────────────────────────────────────

async def _show_question_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    quiz_id = context.user_data.get("dt_quiz_id")
    quiz    = ds.get_quiz(quiz_id)
    questions = quiz.get("questions", [])
    rows = []
    for i, q in enumerate(questions):
        label = q.get("question", "")[:40] or f"سؤال {i + 1}"
        rows.append([InlineKeyboardButton(f"{i + 1}. {label}", callback_data=f"dt_qsel_{i}")])
    rows.append([InlineKeyboardButton("➕ إضافة سؤال", callback_data="dt_q_add_edit")])
    rows.append([InlineKeyboardButton("🔙 رجوع", callback_data="dt_edit_menu")])
    text = f"❓ *أسئلة الاختبار*  —  {len(questions)} سؤال\n\nاختر سؤالاً لتعديله:"
    await _reply(update, text, InlineKeyboardMarkup(rows))
    return DT_E_Q_LIST


async def dt_e_questions(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    return await _show_question_list(update, context)


async def dt_q_add_edit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    context.user_data["dt_mode"] = "edit_add"
    await _reply(update, "📝 أرسل *نص السؤال* الجديد:")
    return DT_C_Q_TEXT


def _q_field_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📝 نص السؤال", callback_data="dt_qf_text")],
        [InlineKeyboardButton("أ) الخيار الأول", callback_data="dt_qf_opta"),
         InlineKeyboardButton("ب) الخيار الثاني", callback_data="dt_qf_optb")],
        [InlineKeyboardButton("✅ الإجابة الصحيحة", callback_data="dt_qf_correct")],
        [InlineKeyboardButton("🗑 حذف السؤال", callback_data="dt_qf_delete")],
        [InlineKeyboardButton("🔙 رجوع لقائمة الأسئلة", callback_data="dt_back_to_q_list")],
    ])


def _q_summary(quiz: dict, idx: int) -> str:
    q = quiz.get("questions", [])[idx]
    correct = "أ" if q.get("correct") == "a" else "ب"
    return (
        f"*السؤال {idx + 1}*\n\n"
        f"{q.get('question', '')}\n\n"
        f"أ) {q.get('option_a', '')}\n"
        f"ب) {q.get('option_b', '')}\n\n"
        f"✅ الإجابة الصحيحة: *{correct}*"
    )


async def dt_qsel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    idx = int(update.callback_query.data.replace("dt_qsel_", ""))
    context.user_data["dt_q_index"] = idx
    quiz_id = context.user_data.get("dt_quiz_id")
    quiz    = ds.get_quiz(quiz_id)
    if idx >= len(quiz.get("questions", [])):
        return await _show_question_list(update, context)
    await _reply(update, _q_summary(quiz, idx) + "\n\nاختر ما تريد تعديله:", _q_field_kb())
    return DT_E_Q_FIELD_MENU


async def dt_back_to_q_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    return await _show_question_list(update, context)


async def _show_q_field_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    quiz_id = context.user_data.get("dt_quiz_id")
    idx     = context.user_data.get("dt_q_index", 0)
    quiz    = ds.get_quiz(quiz_id)
    if idx >= len(quiz.get("questions", [])):
        return await _show_question_list(update, context)
    await _reply(update, _q_summary(quiz, idx) + "\n\nاختر ما تريد تعديله:", _q_field_kb())
    return DT_E_Q_FIELD_MENU


async def dt_qf_text(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    context.user_data["dt_qf_pending"] = "question"
    await _reply(update, "📝 أرسل *نص السؤال* الجديد:")
    return DT_E_Q_FIELD_VAL


async def dt_qf_opta(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    context.user_data["dt_qf_pending"] = "option_a"
    await _reply(update, "أ) أرسل *نص الخيار الأول* الجديد:")
    return DT_E_Q_FIELD_VAL


async def dt_qf_optb(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    context.user_data["dt_qf_pending"] = "option_b"
    await _reply(update, "ب) أرسل *نص الخيار الثاني* الجديد:")
    return DT_E_Q_FIELD_VAL


async def dt_qf_val(update: Update, context: ContextTypes.DEFAULT_TYPE):
    field   = context.user_data.get("dt_qf_pending")
    quiz_id = context.user_data.get("dt_quiz_id")
    idx     = context.user_data.get("dt_q_index", 0)
    text    = update.message.text.strip()
    if field and text:
        ds.update_question(quiz_id, idx, **{field: text})
    await update.message.reply_text("✅ تم التحديث.")
    quiz = ds.get_quiz(quiz_id)
    if idx >= len(quiz.get("questions", [])):
        await update.message.reply_text("❓ *أسئلة الاختبار*", parse_mode=ParseMode.MARKDOWN,
                                         reply_markup=InlineKeyboardMarkup(
                                             [[InlineKeyboardButton("🔙 رجوع", callback_data="dt_edit_menu")]]))
        return DT_E_Q_LIST
    await update.message.reply_text(
        _q_summary(quiz, idx) + "\n\nاختر ما تريد تعديله:",
        parse_mode=ParseMode.MARKDOWN, reply_markup=_q_field_kb(),
    )
    return DT_E_Q_FIELD_MENU


async def dt_qf_correct(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(update, "اختر *الإجابة الصحيحة*:", InlineKeyboardMarkup([[
        InlineKeyboardButton("✅ أ", callback_data="dt_qf_setcorrect_a"),
        InlineKeyboardButton("✅ ب", callback_data="dt_qf_setcorrect_b"),
    ]]))
    return DT_E_Q_FIELD_MENU


async def dt_qf_setcorrect(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    correct = update.callback_query.data.replace("dt_qf_setcorrect_", "")
    quiz_id = context.user_data.get("dt_quiz_id")
    idx     = context.user_data.get("dt_q_index", 0)
    ds.update_question(quiz_id, idx, correct=correct)
    return await _show_q_field_menu(update, context)


async def dt_qf_delete(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(update, "🗑 هل تريد حذف هذا السؤال؟",
                 _yn("dt_qf_delete_yes", "dt_back_to_q_field"))
    return DT_E_Q_DEL_CONFIRM


async def dt_back_to_q_field(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    return await _show_q_field_menu(update, context)


async def dt_qf_delete_yes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    quiz_id = context.user_data.get("dt_quiz_id")
    idx     = context.user_data.get("dt_q_index", 0)
    ds.delete_question(quiz_id, idx)
    await _reply(update, "✅ تم حذف السؤال.")
    return await _show_question_list(update, context)


# ── Cancel fallback ───────────────────────────────────────────────────────────

async def dt_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text("تم الإلغاء.")
    return ConversationHandler.END


# ── Build the ConversationHandler ─────────────────────────────────────────────

def build_distro_admin_handler() -> ConversationHandler:
    hub_reentry = CallbackQueryHandler(dt_hub, pattern="^dt_hub$")
    list_entry  = CallbackQueryHandler(dt_list_entry, pattern=r"^dt_list_(menu|delete|results|split)$")

    return ConversationHandler(
        entry_points=[CallbackQueryHandler(dt_hub, pattern="^adm_distro$")],
        states={
            DT_HUB: [
                CallbackQueryHandler(dt_create, pattern="^dt_create$"),
                list_entry,
            ],
            DT_LIST: [
                CallbackQueryHandler(dt_pick, pattern=r"^dt_pick_\w+$"),
                list_entry, hub_reentry,
            ],
            DT_QUIZ_MENU: [
                CallbackQueryHandler(dt_edit_menu,          pattern="^dt_edit_menu$"),
                CallbackQueryHandler(dt_toggle_visible,     pattern="^dt_toggle_visible$"),
                CallbackQueryHandler(dt_toggle_entry,       pattern="^dt_toggle_entry$"),
                CallbackQueryHandler(dt_show_results,       pattern="^dt_show_results$"),
                CallbackQueryHandler(dt_goto_split,         pattern="^dt_goto_split$"),
                CallbackQueryHandler(dt_delete_confirm,     pattern="^dt_delete_confirm$"),
                CallbackQueryHandler(dt_back_to_quiz_menu,  pattern="^dt_back_to_quiz_menu$"),
                list_entry, hub_reentry,
            ],
            DT_DEL_CONFIRM: [
                CallbackQueryHandler(dt_delete_yes,        pattern="^dt_delete_yes$"),
                CallbackQueryHandler(dt_back_to_quiz_menu, pattern="^dt_back_to_quiz_menu$"),
                hub_reentry,
            ],
            # ── Create ──
            DT_C_NAME:   [MessageHandler(filters.TEXT & ~filters.COMMAND, dt_c_name), hub_reentry],
            DT_C_DESC:   [MessageHandler(filters.TEXT & ~filters.COMMAND, dt_c_desc), hub_reentry],
            DT_C_POINTS: [MessageHandler(filters.TEXT & ~filters.COMMAND, dt_c_points), hub_reentry],
            DT_C_TIMED: [
                CallbackQueryHandler(dt_c_timed_yes, pattern="^dt_timed_yes$"),
                CallbackQueryHandler(dt_c_timed_no,  pattern="^dt_timed_no$"),
                hub_reentry,
            ],
            DT_C_MINUTES: [MessageHandler(filters.TEXT & ~filters.COMMAND, dt_c_minutes), hub_reentry],
            DT_C_VISIBLE: [
                CallbackQueryHandler(dt_c_visible_yes, pattern="^dt_vis_yes$"),
                CallbackQueryHandler(dt_c_visible_no,  pattern="^dt_vis_no$"),
                hub_reentry,
            ],
            # ── Add-question loop (shared) ──
            DT_C_Q_TEXT:   [MessageHandler(filters.TEXT & ~filters.COMMAND, dt_q_text), hub_reentry],
            DT_C_Q_OPT_A:  [MessageHandler(filters.TEXT & ~filters.COMMAND, dt_q_opt_a), hub_reentry],
            DT_C_Q_OPT_B:  [MessageHandler(filters.TEXT & ~filters.COMMAND, dt_q_opt_b), hub_reentry],
            DT_C_Q_CORRECT: [
                CallbackQueryHandler(dt_correct_a, pattern="^dt_correct_a$"),
                CallbackQueryHandler(dt_correct_b, pattern="^dt_correct_b$"),
                hub_reentry,
            ],
            DT_C_Q_MORE: [
                CallbackQueryHandler(dt_q_more_yes,  pattern="^dt_q_more_yes$"),
                CallbackQueryHandler(dt_q_more_done, pattern="^dt_q_more_done$"),
                hub_reentry,
            ],
            # ── Edit test ──
            DT_EDIT_MENU: [
                CallbackQueryHandler(dt_e_name,       pattern="^dt_e_name$"),
                CallbackQueryHandler(dt_e_desc,       pattern="^dt_e_desc$"),
                CallbackQueryHandler(dt_e_points,     pattern="^dt_e_points$"),
                CallbackQueryHandler(dt_e_timed,      pattern="^dt_e_timed$"),
                CallbackQueryHandler(dt_e_questions,  pattern="^dt_e_questions$"),
                CallbackQueryHandler(dt_back_to_quiz_menu, pattern="^dt_back_to_quiz_menu$"),
                hub_reentry,
            ],
            DT_E_NAME:   [MessageHandler(filters.TEXT & ~filters.COMMAND, dt_e_name_val), hub_reentry],
            DT_E_DESC:   [MessageHandler(filters.TEXT & ~filters.COMMAND, dt_e_desc_val), hub_reentry],
            DT_E_POINTS: [MessageHandler(filters.TEXT & ~filters.COMMAND, dt_e_points_val), hub_reentry],
            DT_E_TIMED: [
                CallbackQueryHandler(dt_e_timed_yes, pattern="^dt_e_timed_yes$"),
                CallbackQueryHandler(dt_e_timed_no,  pattern="^dt_e_timed_no$"),
                hub_reentry,
            ],
            DT_E_MINUTES: [MessageHandler(filters.TEXT & ~filters.COMMAND, dt_e_minutes_val), hub_reentry],
            DT_E_Q_LIST: [
                CallbackQueryHandler(dt_qsel,       pattern=r"^dt_qsel_\d+$"),
                CallbackQueryHandler(dt_q_add_edit, pattern="^dt_q_add_edit$"),
                CallbackQueryHandler(dt_edit_menu,  pattern="^dt_edit_menu$"),
                hub_reentry,
            ],
            DT_E_Q_FIELD_MENU: [
                CallbackQueryHandler(dt_qf_text,        pattern="^dt_qf_text$"),
                CallbackQueryHandler(dt_qf_opta,         pattern="^dt_qf_opta$"),
                CallbackQueryHandler(dt_qf_optb,         pattern="^dt_qf_optb$"),
                CallbackQueryHandler(dt_qf_correct,      pattern="^dt_qf_correct$"),
                CallbackQueryHandler(dt_qf_setcorrect,   pattern=r"^dt_qf_setcorrect_(a|b)$"),
                CallbackQueryHandler(dt_qf_delete,       pattern="^dt_qf_delete$"),
                CallbackQueryHandler(dt_back_to_q_list,  pattern="^dt_back_to_q_list$"),
                hub_reentry,
            ],
            DT_E_Q_FIELD_VAL: [MessageHandler(filters.TEXT & ~filters.COMMAND, dt_qf_val), hub_reentry],
            DT_E_Q_DEL_CONFIRM: [
                CallbackQueryHandler(dt_qf_delete_yes,   pattern="^dt_qf_delete_yes$"),
                CallbackQueryHandler(dt_back_to_q_field, pattern="^dt_back_to_q_field$"),
                hub_reentry,
            ],
            # ── Team split ──
            DT_SPLIT_SIZE: [MessageHandler(filters.TEXT & ~filters.COMMAND, dt_split_size_val), hub_reentry],
            DT_SPLIT_VIEW: [
                CallbackQueryHandler(dt_export_teams, pattern="^dt_export_teams$"),
                CallbackQueryHandler(dt_resplit,      pattern="^dt_resplit$"),
                hub_reentry,
            ],
        },
        fallbacks=[CommandHandler("cancel", dt_cancel)],
        allow_reentry=True,
        per_message=False,
    )
