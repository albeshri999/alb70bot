# -*- coding: utf-8 -*-
import io
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    ContextTypes, ConversationHandler, CommandHandler,
    MessageHandler, CallbackQueryHandler, filters,
)
from telegram.constants import ParseMode

from storage import (load_users, load_days, save_days, update_user, load_codes, save_codes,
                     load_credit_log, log_credit_action)
from credits import generate_codes, add_credits, get_balance, normalize_code
import admins_store
from utils import default_question, stage_ordinal

logger = logging.getLogger(__name__)

# ── States ────────────────────────────────────────────────────────────────────
(MAIN,
 # New top-level submenus (pure UI reorganization — same underlying handlers)
 PART_STATS_MENU, COMPETITIONS_MENU, BALANCE_MENU,
 # Add new day
 ADD_NAME, ADD_STAGE_Q, ADD_STAGE_A, ADD_FINAL, ADD_LOCKOUT,
 # Day management hub
 DAY_SEL, STAGE_MANAGE,
 # Stage operations on existing day
 STAGE_ADD_Q, STAGE_ADD_A,
 STAGE_EDIT_SEL, STAGE_EDIT_FIELD, STAGE_EDIT_VAL,
 STAGE_DEL_SEL, STAGE_DEL_CONFIRM,
 STAGE_REORDER,
 # Day-level field editing
 EDIT_DAY_NAME, EDIT_DAY_FINAL, EDIT_DAY_LOCKOUT,
 # Delete day
 DEL_DAY_SEL, DEL_CONFIRM,
 # Reset / broadcast
 RESET_CONFIRM,
 BROADCAST_MSG, BROADCAST_CONFIRM,
 # Code management
 CODE_MANAGE, CODE_COUNT, CODE_POINTS,
 # Credit management
 CREDIT_SEARCH, CREDIT_USER_SEL, CREDIT_USER_MENU,
 CREDIT_ADD, CREDIT_REMOVE, CREDIT_REMOVE_REASON, CREDIT_RESET_CONFIRM,
 # Transaction log
 TRANS_LOG, TRANS_USER_SEARCH, TRANS_DELETE_CONFIRM,
 # Participants list / detail
 PART_LIST, PART_DETAIL,
 PART_BAN_REASON, PART_DELETE_CONFIRM,
 PART_RENAME,
 # Leaderboard settings
 LB_MENU, LB_COUNT_CUSTOM,
 # Word-open notifications toggle
 NOTIF_MENU,
 # Competition days open/closed status manager
 DAYS_TOGGLE_LIST, DAYS_TOGGLE_DAY,
 ) = range(50)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _is_admin(uid: int) -> bool:
    return admins_store.is_admin(uid)


def _md(text: str) -> str:
    """Escape MarkdownV1 special characters in user-provided text."""
    for ch in ("_", "*", "`", "[", "]"):
        text = text.replace(ch, f"\\{ch}")
    return text


async def _reply(update: Update, text: str, keyboard=None):
    kw = {"text": text, "parse_mode": ParseMode.MARKDOWN}
    if keyboard:
        kw["reply_markup"] = keyboard
    if update.callback_query:
        try:
            await update.callback_query.message.edit_text(**kw)
        except Exception as _e1:
            logger.warning("_reply edit_text failed: %s", _e1)
            try:
                await update.callback_query.message.reply_text(**kw)
            except Exception as _e2:
                logger.error("_reply reply_text also failed: %s", _e2)
    else:
        await update.message.reply_text(**kw)


def _main_kb(is_owner: bool = True) -> InlineKeyboardMarkup:
    if not is_owner:
        # Restricted admin — only these six sections are visible/reachable.
        return InlineKeyboardMarkup([
            [InlineKeyboardButton("📊 المشاركون والإحصائيات", callback_data="adm_participants_stats")],
            [InlineKeyboardButton("📝 إدارة الاختبارات",       callback_data="adm_quizzes")],
            [InlineKeyboardButton("👥 اختبار توزيع الفرق",     callback_data="adm_distro")],
            [InlineKeyboardButton("💡 إدارة المبادرات",        callback_data="adm_initiatives")],
            [InlineKeyboardButton("💰 إدارة الأرصدة",          callback_data="adm_balance_menu")],
            [InlineKeyboardButton("📢 الإذاعة",                callback_data="adm_broadcast")],
            [InlineKeyboardButton("🏆 لوحة الشرف",             callback_data="adm_leaderboard")],
        ])
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📊 المشاركون والإحصائيات", callback_data="adm_participants_stats")],
        [InlineKeyboardButton("📝 إدارة الاختبارات",       callback_data="adm_quizzes")],
        [InlineKeyboardButton("👥 اختبار توزيع الفرق",     callback_data="adm_distro")],
        [InlineKeyboardButton("💡 إدارة المبادرات",        callback_data="adm_initiatives")],
        [InlineKeyboardButton("🏆 إدارة المسابقات",        callback_data="adm_competitions_menu")],
        [InlineKeyboardButton("💰 إدارة الأرصدة",          callback_data="adm_balance_menu")],
        [InlineKeyboardButton("📢 الإذاعة",                callback_data="adm_broadcast")],
        [InlineKeyboardButton("🏆 لوحة الشرف",             callback_data="adm_leaderboard")],
        [InlineKeyboardButton("👮 إعدادات المشرفين",       callback_data="adm_admins")],
    ])


# ── New top-level submenus (pure reorganization — each button below calls the
# exact same existing handler function as before; nothing about what these
# handlers DO has changed, only where their entry button lives) ─────────────

def _participants_stats_kb(is_owner: bool) -> InlineKeyboardMarkup:
    rows = [[InlineKeyboardButton("👥 المشاركون", callback_data="adm_participants")]]
    if is_owner:
        rows.append([InlineKeyboardButton("🏆 النتائج", callback_data="adm_results")])
    rows.append([InlineKeyboardButton("📈 الإحصائيات", callback_data="adm_stats")])
    if is_owner:
        rows.append([InlineKeyboardButton("📥 تصدير Excel", callback_data="adm_export")])
    rows.append([InlineKeyboardButton("🔙 رجوع", callback_data="adm_main")])
    return InlineKeyboardMarkup(rows)


async def cb_participants_stats_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(
        update, "📊 *المشاركون والإحصائيات*\n\nاختر العملية:",
        _participants_stats_kb(admins_store.is_owner(update.effective_user.id)),
    )
    return PART_STATS_MENU


def _competitions_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("📅 إدارة المسابقات",      callback_data="adm_manage_days")],
        [InlineKeyboardButton("📅 إدارة أيام المسابقة",  callback_data="adm_days_toggle")],
        [InlineKeyboardButton("➕ إضافة يوم",            callback_data="adm_add_day"),
         InlineKeyboardButton("🗑 حذف يوم",              callback_data="adm_delete_day")],
        [InlineKeyboardButton("🔔 إشعارات فتح الكلمات",  callback_data="adm_notif_toggle")],
        [InlineKeyboardButton("🔄 إعادة تعيين النتائج",  callback_data="adm_reset")],
        [InlineKeyboardButton("🔙 رجوع", callback_data="adm_main")],
    ])


async def cb_competitions_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not admins_store.is_owner(update.effective_user.id):
        await update.callback_query.answer("⛔ هذا الخيار متاح فقط لمالك البوت.", show_alert=True)
        await _reply(update, "⚙️ *لوحة تحكم المشرف*", _main_kb(False))
        return MAIN
    await update.callback_query.answer()
    await _reply(update, "🏆 *إدارة المسابقات*\n\nاختر العملية:", _competitions_kb())
    return COMPETITIONS_MENU


def _balance_kb(is_owner: bool) -> InlineKeyboardMarkup:
    rows = []
    if is_owner:
        rows.append([InlineKeyboardButton("💳 إدارة أكواد الشحن", callback_data="adm_codes")])
    rows.append([InlineKeyboardButton("💰 إدارة أرصدة المشاركين", callback_data="adm_credit")])
    if is_owner:
        rows.append([InlineKeyboardButton("📜 سجل حركة الرصيد", callback_data="adm_tlog")])
    rows.append([InlineKeyboardButton("🔙 رجوع", callback_data="adm_main")])
    return InlineKeyboardMarkup(rows)


async def cb_balance_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(
        update, "💰 *إدارة الأرصدة*\n\nاختر العملية:",
        _balance_kb(admins_store.is_owner(update.effective_user.id)),
    )
    return BALANCE_MENU


def _back_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[
        InlineKeyboardButton("⬅️ القائمة الرئيسية", callback_data="adm_main")
    ]])


def _back_mgmt_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([[
        InlineKeyboardButton("⬅️ رجوع", callback_data="back_to_manage")
    ]])


def _stage_manage_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ إضافة سؤال جديد",      callback_data="stage_add")],
        [InlineKeyboardButton("✏️ تعديل سؤال",            callback_data="stage_edit")],
        [InlineKeyboardButton("🗑 حذف سؤال",              callback_data="stage_del")],
        [InlineKeyboardButton("↕ تغيير ترتيب الأسئلة",   callback_data="stage_reorder")],
        [InlineKeyboardButton("📝 اسم اليوم",             callback_data="stage_name"),
         InlineKeyboardButton("🎯 الكلمة النهائية",       callback_data="stage_final")],
        [InlineKeyboardButton("⏱ مدة الإيقاف",            callback_data="stage_lockout")],
        [InlineKeyboardButton("🔙 رجوع لقائمة الأيام",   callback_data="adm_manage_days"),
         InlineKeyboardButton("⬅️ القائمة الرئيسية",     callback_data="adm_main")],
    ])


async def _show_stage_manage(update: Update, day_key: str, context: ContextTypes.DEFAULT_TYPE):
    """Build and display the stage management hub for a day."""
    days   = load_days()
    d      = days.get(day_key, {})
    stages = d.get("stages", [])

    lines = [f"📅 *{d['name']}*  —  {len(stages)} مراحل\n"]
    for i, s in enumerate(stages):
        q = s.get("question", "") or default_question(i)
        a = s.get("answer", "—")
        lines.append(f"*{i+1}. {stage_ordinal(i)}:*  _{q}_  →  `{a}`")

    if stages:
        lines.append("")
    lines.append(f"الكلمة النهائية: `{d.get('final_word', '—')}`")
    lines.append(f"مدة الإيقاف: {d.get('lockout_hours', 3)} ساعات")
    lines.append("\nاختر العملية:")

    await _reply(update, "\n".join(lines), _stage_manage_kb())
    return STAGE_MANAGE


# ── Entry / main menu ─────────────────────────────────────────────────────────

async def admin_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not _is_admin(update.effective_user.id):
        await update.message.reply_text("❌ ليس لديك صلاحية لاستخدام هذا الأمر.")
        return ConversationHandler.END
    await update.message.reply_text(
        "⚙️ *لوحة تحكم المشرف*",
        reply_markup=_main_kb(admins_store.is_owner(update.effective_user.id)),
        parse_mode=ParseMode.MARKDOWN,
    )
    return MAIN


async def cb_main(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    context.user_data.clear()
    await _reply(update, "⚙️ *لوحة تحكم المشرف*",
                 _main_kb(admins_store.is_owner(update.effective_user.id)))
    return MAIN


# ── Participants ──────────────────────────────────────────────────────────────

_PART_PER_PAGE = 20


def _part_status(user: dict) -> str:
    if user.get("banned"):
        return "🚫 محظور"
    if user.get("completed"):
        return "✅ أكمل"
    if user.get("locked_until"):
        return "🔒 موقوف"
    if user.get("state") == "awaiting_password":
        return "⏳ في التقدم"
    return "🟡 مسجل"


def _part_list_kb(users_list: list, page: int) -> InlineKeyboardMarkup:
    start  = page * _PART_PER_PAGE
    shown  = users_list[start:start + _PART_PER_PAGE]
    total  = len(users_list)

    buttons = [
        [InlineKeyboardButton(
            f"👤 {u.get('full_name', '—')[:28]}",
            callback_data=f"part_detail_{uid}"
        )]
        for uid, u in shown
    ]

    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("⬅️ السابق", callback_data=f"part_page_{page - 1}"))
    if start + _PART_PER_PAGE < total:
        nav.append(InlineKeyboardButton("➡️ التالي", callback_data=f"part_page_{page + 1}"))
    if nav:
        buttons.append(nav)

    buttons.append([InlineKeyboardButton("🔙 رجوع", callback_data="adm_main")])
    return InlineKeyboardMarkup(buttons)


async def _show_part_list(update: Update, context: ContextTypes.DEFAULT_TYPE,
                           page: int = 0) -> int:
    context.user_data["part_page"] = page
    users      = load_users()
    days       = load_days()
    users_list = sorted(users.items(),
                        key=lambda x: (x[1].get("full_name") or "").strip().lower())
    total      = len(users_list)
    total_pages = max(1, (total + _PART_PER_PAGE - 1) // _PART_PER_PAGE)
    start      = page * _PART_PER_PAGE

    if not users_list:
        await _reply(update, "لا يوجد مشاركون بعد.", _back_kb())
        return MAIN

    lines = [f"📋 *قائمة المشاركين*\n"
             f"إجمالي المشاركين: *{total}* | صفحة *{page + 1}* من *{total_pages}*\n"]

    for i, (uid, u) in enumerate(users_list[start:start + _PART_PER_PAGE], start + 1):
        uname    = f"@{_md(u['username'])}" if u.get("username") else "—"
        bal      = u.get("credits", 0)
        status   = _part_status(u)
        day_key  = str(u.get("selected_day") or "")
        day_name = _md(days.get(day_key, {}).get("name", "لم يختر") if day_key else "لم يختر")
        lines.append(
            f"*{i}.* {_md(u.get('full_name', '—'))}\n"
            f"   {uname} | `{uid}`\n"
            f"   💳 {bal} نقطة | {status} | 📅 {day_name}"
        )

    await _reply(update, "\n\n".join(lines), _part_list_kb(users_list, page))
    return PART_LIST


async def cb_participants(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    return await _show_part_list(update, context, page=0)


async def part_page_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    page = int(update.callback_query.data[len("part_page_"):])
    return await _show_part_list(update, context, page=page)


async def part_detail_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    return await _render_part_detail(update, context)


async def _render_part_detail(update: Update, context: ContextTypes.DEFAULT_TYPE):
    data = (update.callback_query.data or "") if update.callback_query else ""
    if data.startswith("part_detail_"):
        uid = data[len("part_detail_"):]
        context.user_data["part_detail_uid"] = uid
    else:
        uid = context.user_data.get("part_detail_uid", "")

    from storage import get_user as _get_user
    user = _get_user(int(uid)) if uid else None
    if not user:
        await _reply(update, "❌ لم يتم العثور على المشارك.", _back_kb())
        return PART_LIST

    days     = load_days()
    day_key  = str(user.get("selected_day") or "")
    day_name = _md(days.get(day_key, {}).get("name", "—") if day_key else "لم يختر بعد")
    uname    = f"@{_md(user['username'])}" if user.get("username") else "—"
    balance  = user.get("credits", 0)
    status   = _part_status(user)
    created  = (user.get("created_at") or "—")[:16].replace("T", " ")
    comp_at  = (user.get("completed_at") or "—")[:16].replace("T", " ")
    name     = _md(user.get("full_name", "—"))

    # Counts from transaction log
    hint_count = 0
    code_count = 0
    last_txns  = []
    try:
        from transactions import load_user_txns, format_entry
        txns       = load_user_txns(int(uid))
        hint_count = sum(1 for t in txns if t.get("type") == "hint_purchase")
        code_count = sum(1 for t in txns if t.get("type") == "recharge_code")
        last_txns  = txns[-3:][::-1]
    except Exception as _te:
        logger.warning("part_detail txn load failed: %s", _te)

    completed_day = day_name if user.get("completed") else "لا"

    lines = [
        "👤 *تفاصيل المشارك*\n",
        f"*الاسم:* {name}",
        f"*Telegram ID:* `{uid}`",
        f"*Username:* {uname}",
        f"*الرصيد:* {balance} نقطة",
        f"*الحالة:* {status}",
        f"*تاريخ التسجيل:* {created}",
        f"*اليوم المختار:* {day_name}",
        f"*الأيام المكتملة:* {completed_day}",
        f"*وقت الإكمال:* {comp_at}",
        f"*عدد مرات كشف الحروف:* {hint_count}",
        f"*عدد الأكواد المستخدمة:* {code_count}",
    ]

    if user.get("banned"):
        reason = user.get("ban_reason") or ""
        lines.append(f"\n🚫 *محظور*" + (f"\n*السبب:* {_md(reason)}" if reason else ""))

    if last_txns:
        sep = "\n" + "─" * 18 + "\n"
        txn_text = sep.join(format_entry(t) for t in last_txns)
        lines.append(f"\n📜 *آخر العمليات:*\n{txn_text}")

    ban_btn = (InlineKeyboardButton("✅ إلغاء الحظر", callback_data="part_unban")
               if user.get("banned")
               else InlineKeyboardButton("🚫 حظر المشارك", callback_data="part_ban"))

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("✏️ إعادة تسمية المتسابق", callback_data="part_rename")],
        [ban_btn],
        [InlineKeyboardButton("🗑 حذف المشارك", callback_data="part_del")],
        [InlineKeyboardButton("🔙 رجوع للقائمة", callback_data="part_back_list")],
        [InlineKeyboardButton("🔙 القائمة الرئيسية", callback_data="adm_main")],
    ])
    await _reply(update, "\n".join(lines), kb)
    return PART_DETAIL


async def part_back_to_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    page = context.user_data.get("part_page", 0)
    return await _show_part_list(update, context, page=page)


# ── Participant ban / unban / delete ─────────────────────────────────────────

def _part_uid(context) -> str:
    return str(context.user_data.get("part_detail_uid", ""))


async def part_ban_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    uid = _part_uid(context)
    from storage import get_user as _get_user
    user = _get_user(int(uid)) if uid else None
    if not user:
        await _reply(update, "❌ لم يتم العثور على المشارك.", _back_kb())
        return PART_LIST

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("⏭ حظر بدون سبب", callback_data="part_ban_skip")],
        [InlineKeyboardButton("❌ إلغاء", callback_data="part_ban_cancel")],
    ])
    await _reply(update,
                 f"🚫 *حظر المشارك:* {_md(user.get('full_name', '—'))}\n\n"
                 "أدخل سبب الحظر (اختياري)، أو اضغط «حظر بدون سبب»:",
                 kb)
    return PART_BAN_REASON


async def _apply_ban(uid: str, reason: str, context: ContextTypes.DEFAULT_TYPE):
    update_user(int(uid), banned=True, ban_reason=reason)
    # Notify the participant
    try:
        msg = "🚫 تم حظر حسابك من قبل إدارة المسابقة."
        if reason:
            msg += f"\n\nالسبب: {reason}"
        await context.bot.send_message(chat_id=int(uid), text=msg)
    except Exception as _e:
        logger.warning("could not notify banned user %s: %s", uid, _e)


async def part_ban_reason_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = _part_uid(context)
    if not uid:
        await _reply(update, "❌ حدث خطأ. حاول مجدداً.", _back_kb())
        return MAIN
    reason = (update.message.text or "").strip()
    await _apply_ban(uid, reason, context)
    return await part_detail_handler_refresh(update, context,
                                             prefix="✅ تم حظر المشارك.\n\n")


async def part_ban_skip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    uid = _part_uid(context)
    if not uid:
        await _reply(update, "❌ حدث خطأ. حاول مجدداً.", _back_kb())
        return MAIN
    await _apply_ban(uid, "", context)
    return await _render_part_detail(update, context)


async def part_ban_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    return await _render_part_detail(update, context)


async def part_unban(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer("✅ تم إلغاء الحظر")
    uid = _part_uid(context)
    if uid:
        update_user(int(uid), banned=False, ban_reason="")
        try:
            await context.bot.send_message(
                chat_id=int(uid),
                text="✅ تم إلغاء الحظر عن حسابك. يمكنك الآن استخدام البوت.\n\nاستخدم /start للمتابعة.")
        except Exception as _e:
            logger.warning("could not notify unbanned user %s: %s", uid, _e)
    return await _render_part_detail(update, context)


async def part_detail_handler_refresh(update: Update, context: ContextTypes.DEFAULT_TYPE,
                                      prefix: str = ""):
    """Re-render detail view after a text-message action (no callback_query)."""
    uid = _part_uid(context)
    from storage import get_user as _get_user
    user = _get_user(int(uid)) if uid else None
    if not user:
        await update.message.reply_text("❌ لم يتم العثور على المشارك.")
        return MAIN
    status = _part_status(user)
    reason = user.get("ban_reason") or ""
    lines = [f"{prefix}👤 *{_md(user.get('full_name', '—'))}*",
             f"*الحالة:* {status}"]
    if user.get("banned") and reason:
        lines.append(f"*سبب الحظر:* {_md(reason)}")
    ban_btn = (InlineKeyboardButton("✅ إلغاء الحظر", callback_data="part_unban")
               if user.get("banned")
               else InlineKeyboardButton("🚫 حظر المشارك", callback_data="part_ban"))
    kb = InlineKeyboardMarkup([
        [ban_btn],
        [InlineKeyboardButton("🗑 حذف المشارك", callback_data="part_del")],
        [InlineKeyboardButton("🔙 رجوع للقائمة", callback_data="part_back_list")],
        [InlineKeyboardButton("🔙 القائمة الرئيسية", callback_data="adm_main")],
    ])
    await update.message.reply_text("\n".join(lines), reply_markup=kb,
                                    parse_mode=ParseMode.MARKDOWN)
    return PART_DETAIL


async def part_delete_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    uid = _part_uid(context)
    from storage import get_user as _get_user
    user = _get_user(int(uid)) if uid else None
    if not user:
        await _reply(update, "❌ لم يتم العثور على المشارك.", _back_kb())
        return PART_LIST

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("✅ نعم، احذف", callback_data="part_del_yes")],
        [InlineKeyboardButton("❌ إلغاء",     callback_data="part_del_no")],
    ])
    await _reply(update,
                 f"⚠️ *تحذير:*\n"
                 f"سيتم حذف جميع بيانات هذا المشارك نهائياً.\n\n"
                 f"*الاسم:* {_md(user.get('full_name', '—'))}\n"
                 f"*Telegram ID:* `{uid}`",
                 kb)
    return PART_DELETE_CONFIRM


async def part_delete_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    uid = _part_uid(context)
    if not uid:
        await _reply(update, "❌ حدث خطأ. حاول مجدداً.", _back_kb())
        return MAIN

    from storage import delete_user as _delete_user
    from transactions import delete_user_txns as _delete_txns
    existed = _delete_user(int(uid))
    removed = _delete_txns(int(uid))
    context.user_data.pop("part_detail_uid", None)
    logger.info("participant %s deleted (existed=%s, txns removed=%d)",
                uid, existed, removed)

    await _reply(update,
                 f"🗑 *تم حذف المشارك نهائياً.*\n"
                 f"تم حذف {removed} عملية من السجل.")
    return await _show_part_list(update, context,
                                 page=context.user_data.get("part_page", 0))


async def part_delete_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    return await _render_part_detail(update, context)


# ── Participant rename ────────────────────────────────────────────────────────

async def part_rename_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    uid = _part_uid(context)
    from storage import get_user as _get_user
    user = _get_user(int(uid)) if uid else None
    if not user:
        await _reply(update, "❌ لم يتم العثور على المشارك.", _back_kb())
        return PART_LIST

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("❌ إلغاء", callback_data="part_rename_cancel")],
    ])
    await _reply(update,
                 f"✏️ *إعادة تسمية المتسابق*\n\n"
                 f"الاسم الحالي: *{_md(user.get('full_name', '—'))}*\n\n"
                 "أدخل الاسم الجديد للمشارك:",
                 kb)
    return PART_RENAME


async def part_rename_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    uid = _part_uid(context)
    if not uid:
        await update.message.reply_text("❌ حدث خطأ. حاول مجدداً.")
        return MAIN

    new_name = (update.message.text or "").strip()
    if not new_name:
        await update.message.reply_text("❌ الاسم لا يمكن أن يكون فارغاً. أدخل الاسم الجديد:")
        return PART_RENAME

    from storage import rename_user as _rename_user, append_admin_log
    from transactions import rename_user_txns as _rename_txns
    from datetime import timezone

    old_name = _rename_user(int(uid), new_name)
    _rename_txns(int(uid), new_name)

    now = datetime.now(timezone.utc)
    append_admin_log({
        "action":    "rename_participant",
        "date":      now.strftime("%d/%m/%Y"),
        "time":      now.strftime("%H:%M:%S"),
        "timestamp": now.isoformat(),
        "admin_id":  update.effective_user.id,
        "user_id":   uid,
        "old_name":  old_name,
        "new_name":  new_name,
    })
    logger.info("admin renamed user %s: %r → %r", uid, old_name, new_name)

    kb = InlineKeyboardMarkup([
        [InlineKeyboardButton("🔙 رجوع للملف الشخصي", callback_data=f"part_detail_{uid}")],
        [InlineKeyboardButton("🔙 القائمة الرئيسية", callback_data="adm_main")],
    ])
    await update.message.reply_text(
        f"✅ *تم تحديث اسم المتسابق بنجاح.*\n\n"
        f"الاسم القديم: {_md(old_name)}\n"
        f"الاسم الجديد: *{_md(new_name)}*",
        reply_markup=kb,
        parse_mode=ParseMode.MARKDOWN,
    )
    return PART_DETAIL


async def part_rename_cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    return await _render_part_detail(update, context)


# ── Results ───────────────────────────────────────────────────────────────────

async def cb_results(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not admins_store.is_owner(update.effective_user.id):
        await update.callback_query.answer("⛔ هذا الخيار متاح فقط لمالك البوت.", show_alert=True)
        await _reply(update, "⚙️ *لوحة تحكم المشرف*", _main_kb(False))
        return MAIN
    await update.callback_query.answer()
    users = load_users()
    days  = load_days()
    lines = ["🏆 *النتائج*\n"]
    for key, day_data in sorted(days.items()):
        day_users = [u for u in users.values() if str(u.get("selected_day")) == key]
        completed = [u for u in day_users if u.get("completed")]
        in_prog   = [u for u in day_users if not u.get("completed")]
        n_stages  = len(day_data.get("stages", []))
        lines.append(f"*{day_data['name']}* ({n_stages} مراحل)")
        lines.append(f"✅ أكملوا: {len(completed)}  ⏳ في التقدم: {len(in_prog)}")
        for u in completed:
            lines.append(f"  • {u.get('full_name', '—')}")
        lines.append("")
    no_day = sum(1 for u in users.values() if not u.get("selected_day"))
    lines.append(f"لم يختاروا يوماً بعد: {no_day}")
    await _reply(update, "\n".join(lines), _back_kb())
    return MAIN


# ── Statistics ────────────────────────────────────────────────────────────────

async def cb_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    users  = load_users()
    days   = load_days()
    codes  = load_codes()
    total     = len(users)
    completed = sum(1 for u in users.values() if u.get("completed"))
    locked    = sum(1 for u in users.values() if u.get("locked_until"))
    unused_c  = sum(1 for v in codes.values() if not v)

    lines = ["📊 *الإحصائيات*\n"]
    lines.append(f"👥 إجمالي المشاركين: *{total}*")
    lines.append(f"✅ أكملوا المسابقة: *{completed}*")
    lines.append(f"🔒 موقوفون حالياً: *{locked}*")
    lines.append(f"📅 عدد الأيام: *{len(days)}*")
    lines.append(f"💳 أكواد غير مستخدمة: *{unused_c}*")
    if total:
        lines.append(f"📈 نسبة الإتمام: *{round(completed/total*100)}%*")
    lines.append("\n*تفصيل الأيام:*")
    for key, d in sorted(days.items()):
        du   = [u for u in users.values() if str(u.get("selected_day")) == key]
        done = sum(1 for u in du if u.get("completed"))
        lines.append(f"  {d['name']}: {len(du)} مشارك، {done} أكمل، {len(d.get('stages', []))} مراحل")
    await _reply(update, "\n".join(lines), _back_kb())
    return MAIN


# ── Export Excel ──────────────────────────────────────────────────────────────

async def cb_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not admins_store.is_owner(update.effective_user.id):
        await update.callback_query.answer("⛔ هذا الخيار متاح فقط لمالك البوت.", show_alert=True)
        await _reply(update, "⚙️ *لوحة تحكم المشرف*", _main_kb(False))
        return MAIN
    await update.callback_query.answer("جاري إنشاء الملف…")
    users = load_users()
    days  = load_days()

    wb = Workbook()
    ws = wb.active
    ws.title = "المشاركون"
    ws.sheet_view.rightToLeft = True

    headers = ["#", "الاسم الكامل", "معرف تيليغرام", "رقم المستخدم",
               "اليوم", "المرحلة", "الحالة", "الرصيد", "موقوف حتى"]
    h_fill = PatternFill("solid", fgColor="1F4E79")
    h_font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    for i, (uid, u) in enumerate(users.items(), 1):
        dk   = str(u.get("selected_day") or "")
        dnam = days.get(dk, {}).get("name", "—") if dk else "—"
        if u.get("completed"):      status = "✅ أكمل"
        elif u.get("locked_until"): status = "🔒 موقوف"
        elif u.get("state") == "awaiting_password": status = "⏳ في التقدم"
        else:                       status = "🟡 مسجل"
        ws.append([i, u.get("full_name","—"),
                   f"@{u['username']}" if u.get("username") else "—",
                   uid, dnam, u.get("password_step", 0), status,
                   u.get("credits", 0), u.get("locked_until") or "—"])

    for idx, w in enumerate([5,25,20,18,25,10,15,10,25], 1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    await update.callback_query.message.reply_document(
        document=buf, filename="participants.xlsx",
        caption="📤 تم تصدير بيانات المشاركين.",
    )
    return MAIN


# ── Manage days (day selection hub) ──────────────────────────────────────────

async def cb_manage_days(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not admins_store.is_owner(update.effective_user.id):
        await update.callback_query.answer("⛔ هذا الخيار متاح فقط لمالك البوت.", show_alert=True)
        await _reply(update, "⚙️ *لوحة تحكم المشرف*", _main_kb(False))
        return MAIN
    await update.callback_query.answer()
    days = load_days()
    if not days:
        await _reply(update, "لا توجد أيام مضافة بعد.\nاستخدم ➕ إضافة يوم جديد.", _back_kb())
        return MAIN
    buttons = [
        [InlineKeyboardButton(f"{d['name']} ({len(d.get('stages',[]))} مراحل)",
                              callback_data=f"manage_day_{k}")]
        for k, d in sorted(days.items())
    ]
    buttons.append([InlineKeyboardButton("⬅️ القائمة الرئيسية", callback_data="adm_main")])
    await _reply(update, "📅 *اختر يوماً للإدارة:*", InlineKeyboardMarkup(buttons))
    return DAY_SEL


async def day_manage_sel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    day_key = update.callback_query.data[len("manage_day_"):]
    days    = load_days()
    if day_key not in days:
        await _reply(update, "اليوم غير موجود.", _back_kb())
        return MAIN
    context.user_data["mgmt_day"] = day_key
    return await _show_stage_manage(update, day_key, context)


async def back_to_manage(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    day_key = context.user_data.get("mgmt_day")
    if not day_key:
        return await cb_main(update, context)
    return await _show_stage_manage(update, day_key, context)


# ── Stage: add to existing day ────────────────────────────────────────────────

async def cb_stage_add(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    day_key   = context.user_data.get("mgmt_day")
    stages    = load_days().get(day_key, {}).get("stages", [])
    stage_num = len(stages)
    def_q     = default_question(stage_num)
    await _reply(update,
        f"➕ *إضافة مرحلة جديدة* (المرحلة {stage_num+1})\n\n"
        f"النص الافتراضي: _{def_q}_\n\n"
        f"أدخل نص السؤال أو اضغط 'استخدام الافتراضي':",
        InlineKeyboardMarkup([[
            InlineKeyboardButton("⏭ استخدام الافتراضي", callback_data="stage_add_q_skip"),
            InlineKeyboardButton("⬅️ رجوع",              callback_data="back_to_manage"),
        ]])
    )
    return STAGE_ADD_Q


async def stage_add_q(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["new_stage_q"] = update.message.text.strip()
    await update.message.reply_text("🔑 أدخل الإجابة (كلمة السر):")
    return STAGE_ADD_A


async def stage_add_q_skip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    context.user_data["new_stage_q"] = ""
    await update.callback_query.message.reply_text("🔑 أدخل الإجابة (كلمة السر):")
    return STAGE_ADD_A


async def stage_add_a(update: Update, context: ContextTypes.DEFAULT_TYPE):
    answer  = update.message.text.strip()
    day_key = context.user_data.get("mgmt_day")
    q       = context.user_data.pop("new_stage_q", "")
    days    = load_days()
    if day_key not in days:
        await update.message.reply_text("حدث خطأ. استخدم /admin للمحاولة مجدداً.")
        return MAIN
    stages = days[day_key].get("stages", [])
    stages.append({"question": q, "answer": answer})
    days[day_key]["stages"] = stages
    save_days(days)
    await update.message.reply_text(
        f"✅ تمت إضافة المرحلة *{stage_ordinal(len(stages)-1)}* بنجاح!",
        parse_mode=ParseMode.MARKDOWN,
    )
    return await _show_stage_manage(update, day_key, context)


# ── Stage: edit in existing day ───────────────────────────────────────────────

async def cb_stage_edit(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    day_key = context.user_data.get("mgmt_day")
    stages  = load_days().get(day_key, {}).get("stages", [])
    if not stages:
        await update.callback_query.answer("لا توجد مراحل للتعديل.", show_alert=True)
        return STAGE_MANAGE
    buttons = [
        [InlineKeyboardButton(
            f"المرحلة {stage_ordinal(i)}: {s.get('answer','—')}",
            callback_data=f"edit_stage_{i}")]
        for i, s in enumerate(stages)
    ]
    buttons.append([InlineKeyboardButton("⬅️ رجوع", callback_data="back_to_manage")])
    await _reply(update, "✏️ *اختر المرحلة للتعديل:*", InlineKeyboardMarkup(buttons))
    return STAGE_EDIT_SEL


async def stage_edit_sel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    idx     = int(update.callback_query.data[len("edit_stage_"):])
    day_key = context.user_data.get("mgmt_day")
    context.user_data["edit_stage_idx"] = idx
    stages  = load_days().get(day_key, {}).get("stages", [])
    if idx >= len(stages):
        return await cb_stage_edit(update, context)
    s = stages[idx]
    q = s.get("question", "") or default_question(idx)
    a = s.get("answer", "—")
    await _reply(update,
        f"✏️ *المرحلة {stage_ordinal(idx)}*\n\nالسؤال: _{q}_\nالإجابة: `{a}`\n\nاختر ما تريد تعديله:",
        InlineKeyboardMarkup([
            [InlineKeyboardButton("📝 تعديل السؤال",   callback_data="sf_q"),
             InlineKeyboardButton("🔑 تعديل الإجابة",  callback_data="sf_a")],
            [InlineKeyboardButton("⬅️ رجوع",           callback_data="stage_edit")],
        ])
    )
    return STAGE_EDIT_FIELD


async def stage_edit_field(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    field   = update.callback_query.data   # "sf_q" or "sf_a"
    idx     = context.user_data.get("edit_stage_idx", 0)
    context.user_data["edit_stage_field"] = field
    label   = f"السؤال للمرحلة {stage_ordinal(idx)}" if field == "sf_q" else f"الإجابة للمرحلة {stage_ordinal(idx)}"
    await _reply(update, f"✏️ أدخل النص الجديد لـ *{label}*:")
    return STAGE_EDIT_VAL


async def stage_edit_val(update: Update, context: ContextTypes.DEFAULT_TYPE):
    new_val = update.message.text.strip()
    day_key = context.user_data.get("mgmt_day")
    idx     = context.user_data.get("edit_stage_idx", 0)
    field   = context.user_data.get("edit_stage_field", "sf_q")
    days    = load_days()
    stages  = days.get(day_key, {}).get("stages", [])
    if idx < len(stages):
        if field == "sf_q":
            stages[idx]["question"] = new_val
        else:
            stages[idx]["answer"] = new_val
        days[day_key]["stages"] = stages
        save_days(days)
    await update.message.reply_text("✅ تم التعديل بنجاح!")
    return await _show_stage_manage(update, day_key, context)


# ── Stage: delete in existing day ────────────────────────────────────────────

async def cb_stage_del(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    day_key = context.user_data.get("mgmt_day")
    stages  = load_days().get(day_key, {}).get("stages", [])
    if not stages:
        await update.callback_query.answer("لا توجد مراحل للحذف.", show_alert=True)
        return STAGE_MANAGE
    buttons = [
        [InlineKeyboardButton(
            f"المرحلة {stage_ordinal(i)}: {s.get('answer','—')}",
            callback_data=f"del_stage_{i}")]
        for i, s in enumerate(stages)
    ]
    buttons.append([InlineKeyboardButton("⬅️ رجوع", callback_data="back_to_manage")])
    await _reply(update, "🗑 *اختر المرحلة للحذف:*", InlineKeyboardMarkup(buttons))
    return STAGE_DEL_SEL


async def stage_del_sel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    idx     = int(update.callback_query.data[len("del_stage_"):])
    day_key = context.user_data.get("mgmt_day")
    context.user_data["del_stage_idx"] = idx
    stages  = load_days().get(day_key, {}).get("stages", [])
    if idx >= len(stages):
        return await cb_stage_del(update, context)
    s = stages[idx]
    q = s.get("question", "") or default_question(idx)
    await _reply(update,
        f"🗑 هل تريد حذف *المرحلة {stage_ordinal(idx)}*؟\n\n"
        f"السؤال: _{q}_\nالإجابة: `{s.get('answer','—')}`\n\n⚠️ لا يمكن التراجع!",
        InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ نعم، احذف", callback_data="delstage_yes"),
             InlineKeyboardButton("❌ إلغاء",       callback_data="back_to_manage")],
        ])
    )
    return STAGE_DEL_CONFIRM


async def stage_del_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer("✅ تم الحذف بنجاح!")
    day_key = context.user_data.get("mgmt_day")
    idx     = context.user_data.pop("del_stage_idx", None)
    days    = load_days()
    stages  = days.get(day_key, {}).get("stages", [])
    if idx is not None and 0 <= idx < len(stages):
        stages.pop(idx)
        days[day_key]["stages"] = stages
        save_days(days)
    return await _show_stage_manage(update, day_key, context)


# ── Stage: reorder ────────────────────────────────────────────────────────────

async def cb_stage_reorder(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    day_key = context.user_data.get("mgmt_day")
    stages  = load_days().get(day_key, {}).get("stages", [])
    if len(stages) < 2:
        await update.callback_query.answer("يجب أن يكون هناك مرحلتان على الأقل.", show_alert=True)
        return STAGE_MANAGE
    lines = ["↕ *تغيير ترتيب المراحل*\n\nالترتيب الحالي:"]
    for i, s in enumerate(stages):
        lines.append(f"{i+1}. {s.get('answer','—')}")
    lines.append(f"\nأدخل الترتيب الجديد كأرقام مفصولة بفاصلة:")
    lines.append(f"مثال: `{','.join(str(len(stages)-i) for i in range(len(stages)))}` لعكس الترتيب")
    await _reply(update, "\n".join(lines), _back_mgmt_kb())
    return STAGE_REORDER


async def stage_reorder_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text    = update.message.text.strip().replace("،", ",")
    day_key = context.user_data.get("mgmt_day")
    days    = load_days()
    stages  = days.get(day_key, {}).get("stages", [])
    n       = len(stages)
    try:
        new_order = [int(x.strip()) - 1 for x in text.split(",")]
        if sorted(new_order) != list(range(n)):
            raise ValueError
    except (ValueError, IndexError):
        await update.message.reply_text(
            f"⚠️ أدخل {n} أرقام من 1 إلى {n} مفصولة بفاصلة.\nمثال: `1,3,2`",
            parse_mode=ParseMode.MARKDOWN,
        )
        return STAGE_REORDER
    days[day_key]["stages"] = [stages[i] for i in new_order]
    save_days(days)
    await update.message.reply_text("✅ تم تغيير الترتيب بنجاح!")
    return await _show_stage_manage(update, day_key, context)


# ── Day-level field editing ───────────────────────────────────────────────────

async def cb_stage_edit_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    day_key = context.user_data.get("mgmt_day")
    current = load_days().get(day_key, {}).get("name", "—")
    await _reply(update, f"📝 الاسم الحالي: *{current}*\n\nأدخل الاسم الجديد:", _back_mgmt_kb())
    return EDIT_DAY_NAME


async def edit_day_name_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day_key = context.user_data.get("mgmt_day")
    days    = load_days()
    days[day_key]["name"] = update.message.text.strip()
    save_days(days)
    await update.message.reply_text("✅ تم تحديث اسم اليوم!")
    return await _show_stage_manage(update, day_key, context)


async def cb_stage_edit_final(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    day_key = context.user_data.get("mgmt_day")
    current = load_days().get(day_key, {}).get("final_word", "—")
    await _reply(update, f"🎯 الكلمة النهائية الحالية: *{current}*\n\nأدخل الكلمة الجديدة:", _back_mgmt_kb())
    return EDIT_DAY_FINAL


async def edit_day_final_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day_key = context.user_data.get("mgmt_day")
    days    = load_days()
    days[day_key]["final_word"] = update.message.text.strip()
    save_days(days)
    await update.message.reply_text("✅ تم تحديث الكلمة النهائية!")
    return await _show_stage_manage(update, day_key, context)


async def cb_stage_edit_lockout(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    day_key = context.user_data.get("mgmt_day")
    current = load_days().get(day_key, {}).get("lockout_hours", 3)
    await _reply(update, f"⏱ مدة الإيقاف الحالية: *{current} ساعات*\n\nأدخل المدة الجديدة بالساعات:", _back_mgmt_kb())
    return EDIT_DAY_LOCKOUT


async def edit_day_lockout_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    day_key = context.user_data.get("mgmt_day")
    try:
        hours = int(update.message.text.strip())
        if hours <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("⚠️ أدخل رقماً صحيحاً أكبر من صفر.")
        return EDIT_DAY_LOCKOUT
    days    = load_days()
    days[day_key]["lockout_hours"] = hours
    save_days(days)
    await update.message.reply_text(f"✅ تم تحديث مدة الإيقاف إلى {hours} ساعات!")
    return await _show_stage_manage(update, day_key, context)


# ── Add new day (unlimited stages) ───────────────────────────────────────────

async def cb_add_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not admins_store.is_owner(update.effective_user.id):
        await update.callback_query.answer("⛔ هذا الخيار متاح فقط لمالك البوت.", show_alert=True)
        await _reply(update, "⚙️ *لوحة تحكم المشرف*", _main_kb(False))
        return MAIN
    await update.callback_query.answer()
    context.user_data["new_day"] = {"name": "", "stages": []}
    await _reply(update, "➕ *إضافة يوم جديد*\n\nأدخل اسم اليوم:", _back_kb())
    return ADD_NAME


async def add_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["new_day"]["name"] = update.message.text.strip()
    stage_num = len(context.user_data["new_day"]["stages"])
    def_q     = default_question(stage_num)
    await update.message.reply_text(
        f"📝 *المرحلة {stage_ordinal(stage_num)} — السؤال*\n\nالافتراضي: _{def_q}_\n\nأدخل نص السؤال:",
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("⏭ استخدام الافتراضي", callback_data="addstage_q_skip"),
        ]])
    )
    return ADD_STAGE_Q


async def add_stage_q(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["cur_q"] = update.message.text.strip()
    stage_num = len(context.user_data["new_day"]["stages"])
    await update.message.reply_text(
        f"🔑 *المرحلة {stage_ordinal(stage_num)} — الإجابة*\n\nأدخل كلمة السر:",
        parse_mode=ParseMode.MARKDOWN,
    )
    return ADD_STAGE_A


async def add_stage_q_skip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    context.user_data["cur_q"] = ""
    stage_num = len(context.user_data["new_day"]["stages"])
    await update.callback_query.message.reply_text(
        f"🔑 *المرحلة {stage_ordinal(stage_num)} — الإجابة*\n\nأدخل كلمة السر:",
        parse_mode=ParseMode.MARKDOWN,
    )
    return ADD_STAGE_A


async def add_stage_a(update: Update, context: ContextTypes.DEFAULT_TYPE):
    answer    = update.message.text.strip()
    q         = context.user_data.pop("cur_q", "")
    context.user_data["new_day"]["stages"].append({"question": q, "answer": answer})
    stage_count = len(context.user_data["new_day"]["stages"])
    await update.message.reply_text(
        f"✅ تمت إضافة *{stage_count}* مراحل.\n\nهل تريد إضافة مرحلة أخرى؟",
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("➕ إضافة مرحلة أخرى", callback_data="addstage_more"),
             InlineKeyboardButton("✅ الانتهاء",           callback_data="addstage_done")],
        ])
    )
    return ADD_STAGE_A


async def add_stage_more(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    stage_num = len(context.user_data["new_day"]["stages"])
    def_q     = default_question(stage_num)
    await update.callback_query.message.reply_text(
        f"📝 *المرحلة {stage_ordinal(stage_num)} — السؤال*\n\nالافتراضي: _{def_q}_\n\nأدخل نص السؤال:",
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("⏭ استخدام الافتراضي", callback_data="addstage_q_skip"),
        ]])
    )
    return ADD_STAGE_Q


async def add_stage_done(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    if not context.user_data["new_day"]["stages"]:
        await update.callback_query.answer("يجب إضافة مرحلة واحدة على الأقل!", show_alert=True)
        return ADD_STAGE_A
    await update.callback_query.message.reply_text(
        "🎯 أدخل الكلمة النهائية (تُعرض عند إكمال جميع المراحل):"
    )
    return ADD_FINAL


async def add_final(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["new_day"]["final_word"] = update.message.text.strip()
    await update.message.reply_text(
        "⏱ أدخل مدة الإيقاف بالساعات بعد 3 محاولات خاطئة:",
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("⏭ تخطي (3 ساعات افتراضي)", callback_data="addlockout_skip"),
        ]])
    )
    return ADD_LOCKOUT


async def add_lockout_skip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    context.user_data["new_day"]["lockout_hours"] = 3
    return await _finish_add_day(update, context, via_callback=True)


async def add_lockout(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        hours = int(update.message.text.strip())
        if hours <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("⚠️ أدخل رقماً صحيحاً أكبر من صفر.")
        return ADD_LOCKOUT
    context.user_data["new_day"]["lockout_hours"] = hours
    return await _finish_add_day(update, context, via_callback=False)


async def _finish_add_day(update: Update, context: ContextTypes.DEFAULT_TYPE, via_callback: bool):
    days    = load_days()
    new_key = str(max((int(k) for k in days.keys()), default=0) + 1)
    nd      = context.user_data.pop("new_day")
    days[new_key] = nd
    save_days(days)

    stages = nd["stages"]
    lines  = [f"✅ *تم إضافة اليوم بنجاح!*\n\n*{nd['name']}* — {len(stages)} مراحل\n"]
    for i, s in enumerate(stages):
        q = s.get("question", "") or default_question(i)
        lines.append(f"*{stage_ordinal(i)}:* _{q}_ → `{s['answer']}`")
    lines.append(f"\nالكلمة النهائية: `{nd.get('final_word', '—')}`")
    lines.append(f"مدة الإيقاف: {nd.get('lockout_hours', 3)} ساعات")

    msg = update.callback_query.message if via_callback else update.message
    await msg.reply_text("\n".join(lines), parse_mode=ParseMode.MARKDOWN, reply_markup=_back_kb())
    return MAIN


# ── Delete day ────────────────────────────────────────────────────────────────

async def cb_delete_day(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not admins_store.is_owner(update.effective_user.id):
        await update.callback_query.answer("⛔ هذا الخيار متاح فقط لمالك البوت.", show_alert=True)
        await _reply(update, "⚙️ *لوحة تحكم المشرف*", _main_kb(False))
        return MAIN
    await update.callback_query.answer()
    days = load_days()
    if not days:
        await _reply(update, "لا توجد أيام للحذف.", _back_kb())
        return MAIN
    buttons = [
        [InlineKeyboardButton(d["name"], callback_data=f"delday_{k}")]
        for k, d in sorted(days.items())
    ]
    buttons.append([InlineKeyboardButton("⬅️ القائمة الرئيسية", callback_data="adm_main")])
    await _reply(update, "🗑 *حذف يوم* — اختر اليوم:", InlineKeyboardMarkup(buttons))
    return DEL_DAY_SEL


async def del_day_sel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    day_key = update.callback_query.data[len("delday_"):]
    days    = load_days()
    if day_key not in days:
        await _reply(update, "اليوم غير موجود.", _back_kb())
        return MAIN
    context.user_data["del_day_key"] = day_key
    d      = days[day_key]
    stages = d.get("stages", [])
    await _reply(update,
        f"🗑 هل تريد حذف *{d['name']}* ({len(stages)} مراحل)؟\n\n⚠️ لا يمكن التراجع!",
        InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ نعم، احذف", callback_data="delday_confirm"),
             InlineKeyboardButton("❌ إلغاء",      callback_data="adm_main")],
        ])
    )
    return DEL_CONFIRM


async def del_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    day_key = context.user_data.pop("del_day_key", None)
    days    = load_days()
    if day_key and day_key in days:
        name = days[day_key]["name"]
        del days[day_key]
        save_days(days)
        await _reply(update, f"✅ تم حذف *{name}* بنجاح.", _back_kb())
    else:
        await _reply(update, "لم يتم العثور على اليوم.", _back_kb())
    return MAIN


# ── Reset results ─────────────────────────────────────────────────────────────

async def cb_reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not admins_store.is_owner(update.effective_user.id):
        await update.callback_query.answer("⛔ هذا الخيار متاح فقط لمالك البوت.", show_alert=True)
        await _reply(update, "⚙️ *لوحة تحكم المشرف*", _main_kb(False))
        return MAIN
    await update.callback_query.answer()
    users = load_users()
    await _reply(update,
        f"🔄 *إعادة تعيين النتائج*\n\n"
        f"سيتم إعادة تعيين تقدم جميع المشاركين ({len(users)} مشارك).\n"
        f"الرصيد لن يُمسح.\n\n⚠️ هل أنت متأكد؟",
        InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ نعم، إعادة تعيين الكل", callback_data="resetconfirm_yes"),
             InlineKeyboardButton("❌ إلغاء",                  callback_data="adm_main")],
        ])
    )
    return RESET_CONFIRM


async def reset_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    users = load_users()
    for uid in users:
        update_user(int(uid),
            state="main_menu", selected_day=None,
            password_step=0, password_attempts=0,
            locked_until=None, completed=False,
            hint_step=-1, hint_revealed=[],
        )
    await _reply(update,
        f"✅ تم إعادة تعيين تقدم *{len(users)}* مشارك.\nالرصيد لم يُمسح.",
        _back_kb()
    )
    return MAIN


# ── Broadcast ─────────────────────────────────────────────────────────────────

async def cb_broadcast(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(update, "📢 *إذاعة رسالة*\n\nأدخل الرسالة:", _back_kb())
    return BROADCAST_MSG


async def broadcast_msg_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    msg_text = update.message.text.strip()
    context.user_data["bcast_msg"] = msg_text
    users = load_users()
    await update.message.reply_text(
        f"📢 *معاينة الرسالة:*\n\n{msg_text}\n\n"
        f"سيتم إرسالها إلى *{len(users)}* مشارك. هل تريد المتابعة؟",
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ إرسال",  callback_data="bcastconfirm_yes"),
             InlineKeyboardButton("❌ إلغاء", callback_data="adm_main")],
        ])
    )
    return BROADCAST_CONFIRM


async def broadcast_confirm_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer("جاري الإرسال…")
    msg_text     = context.user_data.pop("bcast_msg", "")
    users        = load_users()
    sent, failed = 0, 0
    for uid in users:
        try:
            await context.bot.send_message(chat_id=int(uid), text=msg_text)
            sent += 1
        except Exception:
            failed += 1
    await _reply(update,
        f"📢 *اكتمل الإرسال!*\n\n✅ أُرسلت إلى: {sent}\n❌ فشل: {failed}",
        _back_kb()
    )
    return MAIN


# ── Code management ───────────────────────────────────────────────────────────

def _codes_menu_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ إنشاء أكواد جديدة", callback_data="codes_create")],
        [InlineKeyboardButton("📄 عرض جميع الأكواد",  callback_data="codes_view")],
        [InlineKeyboardButton("📤 تصدير الأكواد",      callback_data="codes_export")],
        [InlineKeyboardButton("⬅️ القائمة الرئيسية",  callback_data="adm_main")],
    ])


async def cb_codes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not admins_store.is_owner(update.effective_user.id):
        await update.callback_query.answer("⛔ هذا الخيار متاح فقط لمالك البوت.", show_alert=True)
        await _reply(update, "⚙️ *لوحة تحكم المشرف*", _main_kb(False))
        return MAIN
    await update.callback_query.answer()
    codes  = load_codes()
    unused = sum(1 for v in codes.values() if not normalize_code(v)["used"])
    used   = sum(1 for v in codes.values() if normalize_code(v)["used"])
    await _reply(update,
        f"💳 *إدارة أكواد الشحن*\n\n"
        f"✅ غير مستخدمة: *{unused}*\n❌ مستخدمة: *{used}*\nإجمالي: *{len(codes)}*",
        _codes_menu_kb()
    )
    return CODE_MANAGE


async def cb_codes_view(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    codes  = load_codes()
    unused = [(k, normalize_code(v)) for k, v in codes.items() if not normalize_code(v)["used"]]
    used   = [(k, normalize_code(v)) for k, v in codes.items() if normalize_code(v)["used"]]

    lines = ["📄 *أكواد الشحن*\n"]
    lines.append(f"✅ *غير مستخدمة ({len(unused)}):*")
    if unused:
        shown = unused[:50]
        for i in range(0, len(shown), 3):
            grp = shown[i:i+3]
            lines.append("  `" + "   ".join(f"{c}({o['points']})" for c, o in grp) + "`")
        if len(unused) > 50:
            lines.append(f"  ...و {len(unused)-50} أكواد أخرى")
    else:
        lines.append("  لا توجد أكواد غير مستخدمة")

    lines.append(f"\n❌ *مستخدمة ({len(used)}):*")
    if used:
        shown = used[:30]
        for i in range(0, len(shown), 3):
            grp = shown[i:i+3]
            lines.append("  `" + "   ".join(f"{c}({o['points']})" for c, o in grp) + "`")
        if len(used) > 30:
            lines.append(f"  ...و {len(used)-30} أكواد أخرى")
    else:
        lines.append("  لا توجد أكواد مستخدمة")

    await _reply(update, "\n".join(lines),
        InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ إدارة الأكواد", callback_data="adm_codes")]]))
    return CODE_MANAGE


async def cb_codes_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer("جاري إنشاء الملف…")
    codes = load_codes()

    wb = Workbook()
    ws = wb.active
    ws.title = "أكواد الشحن"
    ws.sheet_view.rightToLeft = True

    headers = ["#", "الكود", "النقاط", "مستخدم", "معرف المستخدم", "تاريخ الاستخدام"]
    h_fill = PatternFill("solid", fgColor="1F4E79")
    h_font = Font(bold=True, color="FFFFFF", name="Arial", size=12)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = h_fill
        cell.font = h_font
        cell.alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 22

    for i, (code, v) in enumerate(codes.items(), 1):
        obj      = normalize_code(v)
        used_at  = obj.get("used_at") or ""
        used_at  = used_at[:16].replace("T", " ") if used_at else "—"
        ws.append([
            i, code,
            obj.get("points", 5),
            "✅ نعم" if obj["used"] else "❌ لا",
            str(obj.get("used_by") or "—"),
            used_at,
        ])

    for idx, w in enumerate([5, 15, 10, 12, 20, 22], 1):
        ws.column_dimensions[ws.cell(1, idx).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    unused_n = sum(1 for v in codes.values() if not normalize_code(v)["used"])
    used_n   = sum(1 for v in codes.values() if normalize_code(v)["used"])
    await update.callback_query.message.reply_document(
        document=buf, filename="recharge_codes.xlsx",
        caption=f"📤 {unused_n} غير مستخدمة / {used_n} مستخدمة",
    )
    return CODE_MANAGE


async def cb_codes_create(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(update, "➕ كم عدد الأكواد التي تريد إنشاءها؟ (1–200)",
        InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ إلغاء", callback_data="adm_codes")]]))
    return CODE_COUNT


async def codes_count_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        count = int(update.message.text.strip())
        if not 1 <= count <= 200:
            raise ValueError
    except ValueError:
        await update.message.reply_text("⚠️ أدخل رقماً من 1 إلى 200.")
        return CODE_COUNT
    context.user_data["codes_count"] = count
    await update.message.reply_text(
        f"✅ سيتم إنشاء *{count}* كود.\n\n"
        "2️⃣ كم عدد النقاط التي يمنحها كل كود؟\n"
        "_مثال: 5 ، 10 ، 20 ، 50 ، 100_",
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("⬅️ إلغاء", callback_data="adm_codes")
        ]])
    )
    return CODE_POINTS


async def codes_points_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        pts = int(update.message.text.strip())
        if pts <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("⚠️ أدخل رقماً صحيحاً أكبر من صفر.")
        return CODE_POINTS
    count     = context.user_data.pop("codes_count", 1)
    new_codes = generate_codes(count, pts)
    rows = []
    for i in range(0, len(new_codes), 3):
        rows.append("  `" + "   ".join(new_codes[i:i+3]) + "`")
    await update.message.reply_text(
        f"✅ تم إنشاء *{len(new_codes)}* كود (كل كود = *{pts}* نقطة):\n\n" + "\n".join(rows),
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=InlineKeyboardMarkup([[
            InlineKeyboardButton("⬅️ إدارة الأكواد", callback_data="adm_codes")
        ]])
    )
    return MAIN


# ── Credit management: helpers ───────────────────────────────────────────────

_CREDIT_PER_PAGE = 10


def _credit_user_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("➕ إضافة نقاط",   callback_data="credit_add"),
         InlineKeyboardButton("➖ خصم نقاط",     callback_data="credit_remove")],
        [InlineKeyboardButton("🔄 تصفير الرصيد", callback_data="credit_reset")],
        [InlineKeyboardButton("📜 سجل العمليات", callback_data="credit_log")],
        [InlineKeyboardButton("🔙 رجوع",          callback_data="credit_back")],
    ])


def _get_sorted_users(query: str | None = None) -> list:
    """Return [(uid, user_dict)] sorted alphabetically; filtered by query if given."""
    users = load_users()
    items = list(users.items())
    if query:
        q = query.strip().lower()
        items = [
            (uid, u) for uid, u in items
            if q in (u.get("full_name") or "").lower()
            or q in (u.get("username")  or "").lower()
            or uid == q.strip()
        ]
    items.sort(key=lambda x: (x[1].get("full_name") or "").lower())
    return items


def _credit_list_kb(users_list: list, page: int,
                    search_query: str | None = None) -> InlineKeyboardMarkup:
    start = page * _CREDIT_PER_PAGE
    shown = users_list[start:start + _CREDIT_PER_PAGE]
    total = len(users_list)

    buttons = [
        [InlineKeyboardButton(
            f"👤 {u.get('full_name', '—')} | {u.get('credits', 0)} نقطة",
            callback_data=f"credit_sel_{uid}"
        )]
        for uid, u in shown
    ]

    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("⬅️ السابق", callback_data=f"credit_page_{page - 1}"))
    if start + _CREDIT_PER_PAGE < total:
        nav.append(InlineKeyboardButton("➡️ التالي", callback_data=f"credit_page_{page + 1}"))
    if nav:
        buttons.append(nav)

    if search_query:
        buttons.append([InlineKeyboardButton("❌ إلغاء البحث", callback_data="credit_list_reset")])
    else:
        buttons.append([InlineKeyboardButton("🔍 بحث", callback_data="credit_search_mode")])

    buttons.append([InlineKeyboardButton("⬅️ القائمة الرئيسية", callback_data="adm_main")])
    return InlineKeyboardMarkup(buttons)


async def _show_credit_list(update: Update, context: ContextTypes.DEFAULT_TYPE,
                             page: int | None = None) -> int:
    search_query = context.user_data.get("credit_search_query")
    if page is None:
        page = context.user_data.get("credit_page", 0)
    context.user_data["credit_page"] = page

    users_list = _get_sorted_users(search_query)
    total      = len(users_list)

    if search_query:
        header = (
            f"🔍 *نتائج البحث:* _{search_query}_\n*{total}* مشارك — اختر مشاركاً:"
            if total else
            f"🔍 لا توجد نتائج للبحث عن: _{search_query}_"
        )
    else:
        header = f"💳 *إدارة أرصدة المشاركين*\n*{total}* مشارك — اختر مشاركاً:"

    await _reply(update, header, _credit_list_kb(users_list, page, search_query))
    return CREDIT_SEARCH


async def _show_credit_user(update: Update, uid: str,
                             context: ContextTypes.DEFAULT_TYPE, prefix: str = "") -> int:
    from storage import get_user as _get_user
    user     = _get_user(int(uid))
    name     = user.get("full_name")  or "—"
    username = f"@{user['username']}" if user.get("username") else "—"
    balance  = user.get("credits", 0)

    created = user.get("created_at") or "—"
    created = created[:16].replace("T", " ") if len(created) >= 16 else created

    day_key = str(user.get("selected_day") or "")
    if day_key:
        days     = load_days()
        day_name = days.get(day_key, {}).get("name", f"يوم {day_key}")
    else:
        day_name = "لم يدخل بعد"

    context.user_data["credit_uid"] = uid
    lines = []
    if prefix:
        lines.append(f"{prefix}\n")
    lines += [
        f"👤 *الاسم:* {name}",
        f"🆔 *User ID:* `{uid}`",
        f"💬 *المعرّف:* {username}",
        f"💰 *الرصيد الحالي:* *{balance}* نقطة",
        f"📅 *تاريخ التسجيل:* {created}",
        f"🏁 *آخر يوم وصل إليه:* {day_name}",
        "\nاختر العملية:",
    ]
    await _reply(update, "\n".join(lines), _credit_user_kb())
    return CREDIT_USER_MENU


# ── Credit management: handlers ───────────────────────────────────────────────

async def cb_credit_manage(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    context.user_data["credit_page"]         = 0
    context.user_data["credit_search_query"] = None
    return await _show_credit_list(update, context, page=0)


async def credit_page_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    page = int(update.callback_query.data[len("credit_page_"):])
    return await _show_credit_list(update, context, page=page)


async def credit_search_mode_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(update,
        "🔍 *بحث عن مشارك*\n\nأدخل الاسم أو Telegram User ID:",
        InlineKeyboardMarkup([[
            InlineKeyboardButton("❌ إلغاء", callback_data="credit_cancel_search")
        ]])
    )
    return CREDIT_USER_SEL


async def credit_cancel_search(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    context.user_data["credit_search_query"] = None
    return await _show_credit_list(update, context, page=0)


async def credit_search_text_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.message.text.strip()
    context.user_data["credit_search_query"] = query
    context.user_data["credit_page"]         = 0
    users_list = _get_sorted_users(query)

    if len(users_list) == 1:
        return await _show_credit_user(update, users_list[0][0], context)

    total  = len(users_list)
    header = (
        f"🔍 *نتائج البحث:* _{query}_\n*{total}* مشارك — اختر مشاركاً:"
        if total else
        f"🔍 لا توجد نتائج للبحث عن: _{query}_"
    )
    await update.message.reply_text(
        header,
        reply_markup=_credit_list_kb(users_list, 0, query),
        parse_mode=ParseMode.MARKDOWN,
    )
    return CREDIT_SEARCH


async def credit_list_reset_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    context.user_data["credit_search_query"] = None
    return await _show_credit_list(update, context, page=0)


async def credit_user_sel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    uid = update.callback_query.data[len("credit_sel_"):]
    return await _show_credit_user(update, uid, context)


async def cb_credit_add(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(update,
        "➕ *إضافة نقاط*\n\nكم عدد النقاط التي تريد إضافتها؟",
        InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ رجوع", callback_data="credit_back")]])
    )
    return CREDIT_ADD


async def credit_add_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        amount = int(update.message.text.strip())
        if amount <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("⚠️ أدخل رقماً صحيحاً أكبر من صفر.")
        return CREDIT_ADD
    uid        = context.user_data.get("credit_uid")
    from storage import get_user as _gu
    user       = _gu(int(uid))
    bal_before = user.get("credits", 0)
    new_bal    = add_credits(int(uid), amount)
    log_credit_action(update.effective_user.id, uid, user.get("full_name", "—"), "add", amount, new_bal)
    try:
        from transactions import record as _rec
        _rec(int(uid), user.get("full_name", "—"),
             "admin_add", amount, bal_before, new_bal,
             f"إضافة {amount} نقطة بواسطة المشرف")
    except Exception as _e:
        logger.warning("transaction record failed: %s", _e)
    return await _show_credit_user(update, uid, context,
                                   prefix=f"✅ تمت إضافة *{amount}* نقطة.")


async def cb_credit_remove(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    uid     = context.user_data.get("credit_uid")
    from storage import get_user as _gu
    balance = _gu(int(uid)).get("credits", 0)
    await _reply(update,
        f"➖ *خصم نقاط*\n\nالرصيد الحالي: *{balance}* نقطة\n\nكم عدد النقاط التي تريد خصمها؟",
        InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ رجوع", callback_data="credit_back")]])
    )
    return CREDIT_REMOVE


async def credit_remove_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        amount = int(update.message.text.strip())
        if amount <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("⚠️ أدخل رقماً صحيحاً أكبر من صفر.")
        return CREDIT_REMOVE
    context.user_data["credit_remove_amount"] = amount
    await update.message.reply_text(
        "📝 أدخل سبب الخصم (سيظهر للمتسابق)،\nأو اضغط «تخطي» للخصم بدون سبب:",
        reply_markup=InlineKeyboardMarkup([
            [InlineKeyboardButton("⏭ تخطي", callback_data="credit_remove_skip")],
            [InlineKeyboardButton("⬅️ إلغاء", callback_data="credit_back")],
        ]),
    )
    return CREDIT_REMOVE_REASON


async def _do_credit_remove(update: Update, context: ContextTypes.DEFAULT_TYPE,
                            reason: str):
    """Perform the deduction, log it, and notify the player.
    The deduction always applies in full, regardless of the current balance —
    the balance is allowed to go negative (no floor at zero)."""
    uid    = context.user_data.get("credit_uid")
    amount = context.user_data.pop("credit_remove_amount", 0)
    from storage import get_user as _gu
    user    = _gu(int(uid))
    balance = user.get("credits", 0)
    deducted = amount
    new_bal  = balance - amount
    update_user(int(uid), credits=new_bal)
    log_credit_action(update.effective_user.id, uid, user.get("full_name", "—"), "remove", deducted, new_bal)
    try:
        from transactions import record as _rec
        desc = f"خصم {deducted} نقطة بواسطة المشرف"
        if reason:
            desc += f" — السبب: {reason}"
        _rec(int(uid), user.get("full_name", "—"),
             "admin_remove", deducted, balance, new_bal, desc)
    except Exception as _e:
        logger.warning("transaction record failed: %s", _e)

    # Notify the player immediately — always show the true new balance,
    # even if it's now negative.
    try:
        reason_line = f"\n\nالسبب:\n{reason}" if reason else ""
        await context.bot.send_message(
            chat_id=int(uid),
            text=(
                "⚠️ تم خصم نقاط من رصيدك.\n\n"
                f"الخصم:\n{deducted} نقطة"
                f"{reason_line}\n\n"
                f"رصيدك الحالي:\n{new_bal} نقطة"
            ),
        )
    except Exception as _e:
        logger.warning("Player penalty notification failed: %s", _e)

    return await _show_credit_user(update, uid, context,
                                   prefix=f"✅ تم خصم *{deducted}* نقطة.")


async def credit_remove_reason_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    reason = update.message.text.strip()
    return await _do_credit_remove(update, context, reason)


async def credit_remove_skip(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    return await _do_credit_remove(update, context, "")


async def cb_credit_reset(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    uid = context.user_data.get("credit_uid")
    from storage import get_user as _gu
    user    = _gu(int(uid))
    name    = user.get("full_name", "—")
    balance = user.get("credits", 0)
    await _reply(update,
        f"🔄 *تصفير الرصيد*\n\nالمشارك: *{name}*\n"
        f"الرصيد الحالي: *{balance}* نقطة\n\n⚠️ هل أنت متأكد؟",
        InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ نعم، تصفير", callback_data="credit_reset_yes"),
             InlineKeyboardButton("❌ إلغاء",       callback_data="credit_back")],
        ])
    )
    return CREDIT_RESET_CONFIRM


async def credit_reset_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer("✅ تم التصفير!")
    uid = context.user_data.get("credit_uid")
    from storage import get_user as _gu
    user    = _gu(int(uid))
    old_bal = user.get("credits", 0)
    update_user(int(uid), credits=0)
    log_credit_action(update.effective_user.id, uid, user.get("full_name", "—"), "reset", old_bal, 0)
    try:
        from transactions import record as _rec
        _rec(int(uid), user.get("full_name", "—"),
             "admin_reset", old_bal, old_bal, 0,
             "تصفير الرصيد بواسطة المشرف")
    except Exception as _e:
        logger.warning("transaction record failed: %s", _e)
    return await _show_credit_user(update, uid, context, prefix="✅ تم تصفير الرصيد.")


async def cb_credit_log(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    uid   = context.user_data.get("credit_uid")
    from storage import get_user as _gu
    user  = _gu(int(uid))
    name  = user.get("full_name", "—")
    log   = load_credit_log()
    u_log = [e for e in log if str(e.get("user_id")) == str(uid)]

    if not u_log:
        await _reply(update,
            f"📜 *سجل العمليات — {name}*\n\nلا توجد عمليات مسجلة.",
            InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ رجوع", callback_data="credit_back")]])
        )
        return CREDIT_USER_MENU

    labels = {"add": "➕ إضافة", "remove": "➖ خصم", "reset": "🔄 تصفير"}
    lines  = [f"📜 *سجل العمليات — {name}*\n"]
    for e in reversed(u_log[-20:]):
        ts     = (e.get("timestamp") or "")[:16].replace("T", " ")
        action = labels.get(e.get("action", ""), "—")
        amt    = e.get("amount", 0)
        new_b  = e.get("new_balance", 0)
        lines.append(f"`{ts}` | {action} {amt} نقطة | الرصيد: {new_b}")

    await _reply(update, "\n".join(lines),
        InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ رجوع", callback_data="credit_back")]]))
    return CREDIT_USER_MENU


async def credit_back(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    context.user_data.pop("credit_uid", None)
    return await _show_credit_list(update, context)


# ── Transaction log ────────────────────────────────────────────────────────────

_TLOG_PER_PAGE = 8


def _tlog_main_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("1️⃣ عرض سجل مشارك",      callback_data="tlog_user")],
        [InlineKeyboardButton("2️⃣ عرض جميع العمليات",  callback_data="tlog_all")],
        [InlineKeyboardButton("3️⃣ تصدير Excel",         callback_data="tlog_excel")],
        [InlineKeyboardButton("4️⃣ تصدير PDF",           callback_data="tlog_pdf")],
        [InlineKeyboardButton("5️⃣ حذف السجل",           callback_data="tlog_delete")],
        [InlineKeyboardButton("⬅️ القائمة الرئيسية",     callback_data="adm_main")],
    ])


def _tlog_pager(txns: list, page: int, page_cb: str) -> tuple[str, InlineKeyboardMarkup]:
    """Return (text, keyboard) for one page of transactions."""
    from transactions import format_entry
    total  = len(txns)
    start  = page * _TLOG_PER_PAGE
    shown  = txns[start:start + _TLOG_PER_PAGE]
    sep    = "\n\n" + "─" * 20 + "\n\n"
    body   = sep.join(format_entry(t) for t in shown) if shown else "لا توجد عمليات."
    pages  = max(1, (total + _TLOG_PER_PAGE - 1) // _TLOG_PER_PAGE)
    header = f"📄 الصفحة {page + 1} من {pages}  |  *{total}* عملية\n\n"

    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("⬅️ السابق", callback_data=f"{page_cb}_{page - 1}"))
    if start + _TLOG_PER_PAGE < total:
        nav.append(InlineKeyboardButton("➡️ التالي", callback_data=f"{page_cb}_{page + 1}"))
    buttons = []
    if nav:
        buttons.append(nav)
    buttons.append([InlineKeyboardButton("🔙 رجوع", callback_data="adm_tlog")])
    return header + body, InlineKeyboardMarkup(buttons)


async def cb_tlog_main(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not admins_store.is_owner(update.effective_user.id):
        await update.callback_query.answer("⛔ هذا الخيار متاح فقط لمالك البوت.", show_alert=True)
        await _reply(update, "⚙️ *لوحة تحكم المشرف*", _main_kb(False))
        return MAIN
    await update.callback_query.answer()
    from transactions import load_all
    total = len(load_all())
    await _reply(update,
        f"📜 *سجل حركة الرصيد*\n\n*{total}* عملية مسجلة — اختر:",
        _tlog_main_kb(),
    )
    return TRANS_LOG


async def tlog_user_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(update,
        "🔍 *عرض سجل مشارك*\n\nأدخل الاسم أو Telegram User ID:",
        InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ رجوع", callback_data="adm_tlog")]]),
    )
    return TRANS_USER_SEARCH


async def tlog_user_search_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    from transactions import load_user_txns
    query      = update.message.text.strip()
    back_kb    = InlineKeyboardMarkup([[InlineKeyboardButton("⬅️ رجوع", callback_data="adm_tlog")]])

    # Find matching user(s) from users.json
    matched = _get_sorted_users(query)
    uid = None
    if matched:
        uid = matched[0][0]
    elif query.isdigit():
        uid = query

    if not uid:
        await update.message.reply_text("❌ لم يتم العثور على مشارك.", reply_markup=back_kb)
        return TRANS_USER_SEARCH

    txns = load_user_txns(int(uid))
    if not txns:
        name = matched[0][1].get("full_name", uid) if matched else uid
        await update.message.reply_text(
            f"📭 لا توجد عمليات مسجلة للمشارك *{name}*.",
            reply_markup=back_kb,
            parse_mode=ParseMode.MARKDOWN,
        )
        return TRANS_USER_SEARCH

    context.user_data["tlog_uid"] = uid
    name = (matched[0][1].get("full_name") if matched else None) or txns[-1].get("full_name", uid)
    text, kb = _tlog_pager(txns, 0, "tlog_upage")
    await update.message.reply_text(
        f"👤 *{name}*\n\n" + text,
        reply_markup=kb,
        parse_mode=ParseMode.MARKDOWN,
    )
    return TRANS_LOG


async def tlog_upage_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    from transactions import load_user_txns
    from storage import get_user as _gu
    page = int(update.callback_query.data[len("tlog_upage_"):])
    uid  = context.user_data.get("tlog_uid", "")
    txns = load_user_txns(int(uid)) if uid else []
    user = _gu(int(uid)) if uid else {}
    name = user.get("full_name") or (txns[-1].get("full_name") if txns else uid)
    text, kb = _tlog_pager(txns, page, "tlog_upage")
    await _reply(update, f"👤 *{name}*\n\n" + text, kb)
    return TRANS_LOG


async def tlog_all_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    from transactions import load_all
    txns = load_all()
    if not txns:
        await _reply(update, "📭 لا توجد عمليات مسجلة.", _tlog_main_kb())
        return TRANS_LOG
    context.user_data.pop("tlog_uid", None)
    text, kb = _tlog_pager(txns, 0, "tlog_page")
    await _reply(update, text, kb)
    return TRANS_LOG


async def tlog_page_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    from transactions import load_all
    page = int(update.callback_query.data[len("tlog_page_"):])
    txns = load_all()
    text, kb = _tlog_pager(txns, page, "tlog_page")
    await _reply(update, text, kb)
    return TRANS_LOG


async def tlog_excel_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer("⏳ جاري التصدير...")
    from transactions import load_all, export_excel
    txns = load_all()
    if not txns:
        await _reply(update, "📭 لا توجد عمليات للتصدير.", _tlog_main_kb())
        return TRANS_LOG
    buf = export_excel(txns)
    await update.callback_query.message.reply_document(
        document=buf,
        filename="transactions.xlsx",
        caption=f"📊 سجل حركة الرصيد — {len(txns)} عملية",
    )
    return TRANS_LOG


async def tlog_pdf_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer("⏳ جاري التصدير...")
    from transactions import load_all, export_pdf
    txns = load_all()
    if not txns:
        await _reply(update, "📭 لا توجد عمليات للتصدير.", _tlog_main_kb())
        return TRANS_LOG
    buf = export_pdf(txns)
    await update.callback_query.message.reply_document(
        document=buf,
        filename="transactions.pdf",
        caption=f"📄 سجل حركة الرصيد — {len(txns)} عملية",
    )
    return TRANS_LOG


async def tlog_delete_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    from transactions import load_all
    total = len(load_all())
    await _reply(update,
        f"🗑 *حذف سجل الحركات*\n\n"
        f"⚠️ هل أنت متأكد من حذف جميع العمليات (*{total}*)؟\n"
        f"لا يمكن التراجع عن هذا الإجراء.",
        InlineKeyboardMarkup([
            [InlineKeyboardButton("✅ نعم، احذف الكل", callback_data="tlog_delete_yes"),
             InlineKeyboardButton("❌ إلغاء",            callback_data="adm_tlog")],
        ])
    )
    return TRANS_DELETE_CONFIRM


async def tlog_delete_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer("✅ تم الحذف!")
    from transactions import delete_all
    delete_all()
    await _reply(update, "✅ *تم حذف سجل الحركات بالكامل.*", _tlog_main_kb())
    return TRANS_LOG


# ── Leaderboard settings ──────────────────────────────────────────────────────

def _lb_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("👁 إظهار لوحة الشرف", callback_data="lb_show")],
        [InlineKeyboardButton("🙈 إخفاء لوحة الشرف", callback_data="lb_hide")],
        [InlineKeyboardButton("🔢 عدد المتصدرين",    callback_data="lb_count")],
        [InlineKeyboardButton("🔙 رجوع",              callback_data="adm_main")],
    ])


def _lb_status_text() -> str:
    from storage import get_leaderboard_visible, get_leaderboard_count
    visible = get_leaderboard_visible()
    count   = get_leaderboard_count()
    status  = "👁 ظاهرة" if visible else "🙈 مخفية"
    return (
        "🏆 *إعدادات لوحة الشرف*\n\n"
        f"*الحالة:* {status}\n"
        f"*عدد المتصدرين:* {count}"
    )


async def cb_leaderboard(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(update, _lb_status_text(), _lb_kb())
    return LB_MENU


async def cb_lb_show(update: Update, context: ContextTypes.DEFAULT_TYPE):
    from storage import set_leaderboard_visible
    set_leaderboard_visible(True)
    await update.callback_query.answer("✅ تم إظهار لوحة الشرف.")
    await _reply(update, "✅ تم إظهار لوحة الشرف.\n\n" + _lb_status_text(), _lb_kb())
    return LB_MENU


async def cb_lb_hide(update: Update, context: ContextTypes.DEFAULT_TYPE):
    from storage import set_leaderboard_visible
    set_leaderboard_visible(False)
    await update.callback_query.answer("✅ تم إخفاء لوحة الشرف.")
    await _reply(update, "✅ تم إخفاء لوحة الشرف.\n\n" + _lb_status_text(), _lb_kb())
    return LB_MENU


async def cb_lb_count(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    from storage import get_leaderboard_count
    current = get_leaderboard_count()
    await _reply(update,
        f"🔢 *عدد المتصدرين*\n\nالعدد الحالي: *{current}*\n\nاختر العدد المطلوب إظهاره:",
        InlineKeyboardMarkup([
            [InlineKeyboardButton("3",  callback_data="lb_count_3"),
             InlineKeyboardButton("5",  callback_data="lb_count_5")],
            [InlineKeyboardButton("10", callback_data="lb_count_10"),
             InlineKeyboardButton("20", callback_data="lb_count_20")],
            [InlineKeyboardButton("✏️ عدد مخصص", callback_data="lb_count_custom")],
            [InlineKeyboardButton("🔙 رجوع",      callback_data="adm_leaderboard")],
        ])
    )
    return LB_MENU


async def cb_lb_count_set(update: Update, context: ContextTypes.DEFAULT_TYPE):
    count = int(update.callback_query.data[len("lb_count_"):])
    from storage import set_leaderboard_count
    set_leaderboard_count(count)
    await update.callback_query.answer(f"✅ تم ضبط العدد على {count}.")
    await _reply(update, f"✅ تم ضبط عدد المتصدرين على *{count}*.\n\n" + _lb_status_text(),
                 _lb_kb())
    return LB_MENU


async def cb_lb_count_custom(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    await _reply(update,
        "✏️ أرسل عدد المتصدرين المطلوب إظهاره.",
        InlineKeyboardMarkup([[
            InlineKeyboardButton("🔙 رجوع", callback_data="adm_leaderboard")
        ]])
    )
    return LB_COUNT_CUSTOM


async def lb_count_custom_handler(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        count = int(update.message.text.strip())
        if count <= 0:
            raise ValueError
    except ValueError:
        await update.message.reply_text("⚠️ أدخل رقماً صحيحاً أكبر من صفر.")
        return LB_COUNT_CUSTOM
    from storage import set_leaderboard_count
    set_leaderboard_count(count)
    await update.message.reply_text(
        f"✅ تم ضبط عدد المتصدرين على *{count}*.\n\n" + _lb_status_text(),
        parse_mode=ParseMode.MARKDOWN,
        reply_markup=_lb_kb(),
    )
    return LB_MENU


# ── Word-open notifications toggle ────────────────────────────────────────────

def _notif_status_text() -> str:
    from storage import get_notify_word_open
    enabled = get_notify_word_open()
    status  = "✅ مفعلة" if enabled else "❌ معطلة"
    return f"🔔 *إشعارات فتح الكلمات*\n\nالحالة الحالية: {status}"


def _notif_kb() -> InlineKeyboardMarkup:
    from storage import get_notify_word_open
    enabled = get_notify_word_open()
    label   = "⏸ إيقاف" if enabled else "▶️ تفعيل"
    cb      = "notif_disable" if enabled else "notif_enable"
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(label, callback_data=cb)],
        [InlineKeyboardButton("⬅️ القائمة الرئيسية", callback_data="adm_main")],
    ])


async def cb_notif_menu(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not admins_store.is_owner(update.effective_user.id):
        await update.callback_query.answer("⛔ هذا الخيار متاح فقط لمالك البوت.", show_alert=True)
        await _reply(update, "⚙️ *لوحة تحكم المشرف*", _main_kb(False))
        return MAIN
    await update.callback_query.answer()
    await _reply(update, _notif_status_text(), _notif_kb())
    return NOTIF_MENU


async def cb_notif_enable(update: Update, context: ContextTypes.DEFAULT_TYPE):
    from storage import set_notify_word_open
    set_notify_word_open(True)
    await update.callback_query.answer("✅ تم تفعيل الإشعارات.")
    await _reply(update,
        "✅ تم تفعيل إشعارات فتح الكلمات.\n\n" + _notif_status_text(),
        _notif_kb())
    return NOTIF_MENU


async def cb_notif_disable(update: Update, context: ContextTypes.DEFAULT_TYPE):
    from storage import set_notify_word_open
    set_notify_word_open(False)
    await update.callback_query.answer("✅ تم إيقاف الإشعارات.")
    await _reply(update,
        "✅ تم إيقاف إشعارات فتح الكلمات.\n\n" + _notif_status_text(),
        _notif_kb())
    return NOTIF_MENU


# ── Competition days open/closed status manager ───────────────────────────────

def _days_toggle_list_kb() -> InlineKeyboardMarkup:
    from storage import get_day_open
    days = load_days()
    buttons = []
    for k, d in sorted(days.items()):
        icon = "🟢" if get_day_open(k) else "🔴"
        buttons.append([InlineKeyboardButton(
            f"{icon} {d.get('name', f'اليوم {k}')}",
            callback_data=f"daytoggle_sel_{k}",
        )])
    buttons.append([InlineKeyboardButton("⬅️ القائمة الرئيسية", callback_data="adm_main")])
    return InlineKeyboardMarkup(buttons)


async def cb_days_toggle_list(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not admins_store.is_owner(update.effective_user.id):
        await update.callback_query.answer("⛔ هذا الخيار متاح فقط لمالك البوت.", show_alert=True)
        await _reply(update, "⚙️ *لوحة تحكم المشرف*", _main_kb(False))
        return MAIN
    await update.callback_query.answer()
    days = load_days()
    if not days:
        await _reply(update, "لا توجد أيام مضافة بعد.\nاستخدم ➕ إضافة يوم جديد.", _back_kb())
        return MAIN
    await _reply(update,
        "📅 *إدارة أيام المسابقة*\n\nاختر يوماً لعرض حالته والتحكم به:",
        _days_toggle_list_kb())
    return DAYS_TOGGLE_LIST


def _day_toggle_status_text(day_key: str) -> str:
    from storage import get_day_open
    d      = load_days().get(day_key, {})
    name   = d.get("name", f"اليوم {day_key}")
    status = "🟢 مفتوح للمشاركة" if get_day_open(day_key) else "🔴 مغلق"
    return f"📅 *{name}*\n\nالحالة الحالية: {status}"


def _day_toggle_kb() -> InlineKeyboardMarkup:
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🟢 تفعيل اليوم", callback_data="daytoggle_open"),
         InlineKeyboardButton("🔴 إيقاف اليوم", callback_data="daytoggle_close")],
        [InlineKeyboardButton("🔙 رجوع لقائمة الأيام", callback_data="adm_days_toggle")],
        [InlineKeyboardButton("⬅️ القائمة الرئيسية",   callback_data="adm_main")],
    ])


async def day_toggle_sel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.callback_query.answer()
    day_key = update.callback_query.data[len("daytoggle_sel_"):]
    days    = load_days()
    if day_key not in days:
        await _reply(update, "اليوم غير موجود.", _back_kb())
        return MAIN
    context.user_data["toggle_day"] = day_key
    await _reply(update, _day_toggle_status_text(day_key), _day_toggle_kb())
    return DAYS_TOGGLE_DAY


async def day_toggle_open(update: Update, context: ContextTypes.DEFAULT_TYPE):
    from storage import set_day_open
    day_key = context.user_data.get("toggle_day")
    if not day_key or day_key not in load_days():
        return await cb_days_toggle_list(update, context)
    set_day_open(day_key, True)
    await update.callback_query.answer("✅ تم فتح اليوم.")
    await _reply(update,
        "✅ تم فتح هذا اليوم للمشاركة.\n\n" + _day_toggle_status_text(day_key),
        _day_toggle_kb())
    return DAYS_TOGGLE_DAY


async def day_toggle_close(update: Update, context: ContextTypes.DEFAULT_TYPE):
    from storage import set_day_open
    day_key = context.user_data.get("toggle_day")
    if not day_key or day_key not in load_days():
        return await cb_days_toggle_list(update, context)
    set_day_open(day_key, False)
    await update.callback_query.answer("✅ تم إيقاف اليوم.")
    await _reply(update,
        "✅ تم إيقاف هذا اليوم عن المشاركة.\n\n" + _day_toggle_status_text(day_key),
        _day_toggle_kb())
    return DAYS_TOGGLE_DAY


# ── Cancel ────────────────────────────────────────────────────────────────────

async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data.clear()
    await update.message.reply_text(
        "تم إلغاء العملية. استخدم /admin للعودة."
    )
    return ConversationHandler.END


# ── Build handler ─────────────────────────────────────────────────────────────

def build_admin_handler() -> ConversationHandler:
    back      = CallbackQueryHandler(cb_main,        pattern="^adm_main$")
    back_mgmt = CallbackQueryHandler(back_to_manage, pattern="^back_to_manage$")

    return ConversationHandler(
        entry_points=[CommandHandler("admin", admin_start)],
        states={
            MAIN: [
                back,
                CallbackQueryHandler(cb_participants_stats_menu, pattern="^adm_participants_stats$"),
                CallbackQueryHandler(cb_competitions_menu,        pattern="^adm_competitions_menu$"),
                CallbackQueryHandler(cb_balance_menu,             pattern="^adm_balance_menu$"),
                CallbackQueryHandler(cb_broadcast,     pattern="^adm_broadcast$"),
                CallbackQueryHandler(cb_leaderboard,   pattern="^adm_leaderboard$"),
                # NOTE: adm_quizzes / adm_distro / adm_initiatives / adm_admins are
                # intentionally NOT handled here — they fall through untouched to
                # their own independent ConversationHandlers (quiz_admin.py /
                # distro_admin.py / initiatives_admin.py / admin_settings.py),
                # exactly as before this reorganization.
            ],
            # ── Participants & stats submenu ("📊 المشاركون والإحصائيات")
            PART_STATS_MENU: [
                CallbackQueryHandler(cb_participants, pattern="^adm_participants$"),
                CallbackQueryHandler(cb_results,      pattern="^adm_results$"),
                CallbackQueryHandler(cb_stats,        pattern="^adm_stats$"),
                CallbackQueryHandler(cb_export,       pattern="^adm_export$"),
                back,
            ],
            # ── Competition management submenu ("🏆 إدارة المسابقات")
            COMPETITIONS_MENU: [
                CallbackQueryHandler(cb_manage_days,      pattern="^adm_manage_days$"),
                CallbackQueryHandler(cb_days_toggle_list, pattern="^adm_days_toggle$"),
                CallbackQueryHandler(cb_add_day,          pattern="^adm_add_day$"),
                CallbackQueryHandler(cb_delete_day,       pattern="^adm_delete_day$"),
                CallbackQueryHandler(cb_notif_menu,       pattern="^adm_notif_toggle$"),
                CallbackQueryHandler(cb_reset,            pattern="^adm_reset$"),
                back,
            ],
            # ── Balance management submenu ("💰 إدارة الأرصدة")
            BALANCE_MENU: [
                CallbackQueryHandler(cb_codes,         pattern="^adm_codes$"),
                CallbackQueryHandler(cb_credit_manage, pattern="^adm_credit$"),
                CallbackQueryHandler(cb_tlog_main,     pattern="^adm_tlog$"),
                back,
            ],
            # ── Word-open notifications toggle
            NOTIF_MENU: [
                CallbackQueryHandler(cb_notif_enable,  pattern="^notif_enable$"),
                CallbackQueryHandler(cb_notif_disable, pattern="^notif_disable$"),
                back,
            ],
            # ── Competition days open/closed status manager
            DAYS_TOGGLE_LIST: [
                CallbackQueryHandler(day_toggle_sel, pattern=r"^daytoggle_sel_.+$"),
                back,
            ],
            DAYS_TOGGLE_DAY: [
                CallbackQueryHandler(day_toggle_open,      pattern="^daytoggle_open$"),
                CallbackQueryHandler(day_toggle_close,     pattern="^daytoggle_close$"),
                CallbackQueryHandler(cb_days_toggle_list,  pattern="^adm_days_toggle$"),
                back,
            ],
            # ── Leaderboard settings
            LB_MENU: [
                CallbackQueryHandler(cb_lb_show,         pattern="^lb_show$"),
                CallbackQueryHandler(cb_lb_hide,         pattern="^lb_hide$"),
                CallbackQueryHandler(cb_lb_count,        pattern="^lb_count$"),
                CallbackQueryHandler(cb_lb_count_set,    pattern=r"^lb_count_\d+$"),
                CallbackQueryHandler(cb_lb_count_custom, pattern="^lb_count_custom$"),
                CallbackQueryHandler(cb_leaderboard,     pattern="^adm_leaderboard$"),
                back,
            ],
            LB_COUNT_CUSTOM: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, lb_count_custom_handler),
                CallbackQueryHandler(cb_leaderboard, pattern="^adm_leaderboard$"),
                back,
            ],
            # ── Add new day
            ADD_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, add_name), back,
            ],
            ADD_STAGE_Q: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, add_stage_q),
                CallbackQueryHandler(add_stage_q_skip, pattern="^addstage_q_skip$"),
                back,
            ],
            ADD_STAGE_A: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, add_stage_a),
                CallbackQueryHandler(add_stage_more,  pattern="^addstage_more$"),
                CallbackQueryHandler(add_stage_done,  pattern="^addstage_done$"),
                back,
            ],
            ADD_FINAL: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, add_final), back,
            ],
            ADD_LOCKOUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, add_lockout),
                CallbackQueryHandler(add_lockout_skip, pattern="^addlockout_skip$"),
                back,
            ],
            # ── Day management hub
            DAY_SEL: [
                CallbackQueryHandler(day_manage_sel, pattern=r"^manage_day_\d+$"),
                back,
            ],
            STAGE_MANAGE: [
                CallbackQueryHandler(cb_stage_add,        pattern="^stage_add$"),
                CallbackQueryHandler(cb_stage_edit,       pattern="^stage_edit$"),
                CallbackQueryHandler(cb_stage_del,        pattern="^stage_del$"),
                CallbackQueryHandler(cb_stage_reorder,    pattern="^stage_reorder$"),
                CallbackQueryHandler(cb_stage_edit_name,   pattern="^stage_name$"),
                CallbackQueryHandler(cb_stage_edit_final,  pattern="^stage_final$"),
                CallbackQueryHandler(cb_stage_edit_lockout,pattern="^stage_lockout$"),
                CallbackQueryHandler(cb_manage_days,       pattern="^adm_manage_days$"),
                back,
            ],
            # ── Stage add
            STAGE_ADD_Q: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, stage_add_q),
                CallbackQueryHandler(stage_add_q_skip, pattern="^stage_add_q_skip$"),
                back_mgmt, back,
            ],
            STAGE_ADD_A: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, stage_add_a),
                back_mgmt, back,
            ],
            # ── Stage edit
            STAGE_EDIT_SEL: [
                CallbackQueryHandler(stage_edit_sel, pattern=r"^edit_stage_\d+$"),
                back_mgmt, back,
            ],
            STAGE_EDIT_FIELD: [
                CallbackQueryHandler(stage_edit_field, pattern="^sf_[qa]$"),
                CallbackQueryHandler(cb_stage_edit,    pattern="^stage_edit$"),
                back_mgmt, back,
            ],
            STAGE_EDIT_VAL: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, stage_edit_val),
                back_mgmt, back,
            ],
            # ── Stage delete
            STAGE_DEL_SEL: [
                CallbackQueryHandler(stage_del_sel, pattern=r"^del_stage_\d+$"),
                back_mgmt, back,
            ],
            STAGE_DEL_CONFIRM: [
                CallbackQueryHandler(stage_del_confirm, pattern="^delstage_yes$"),
                back_mgmt, back,
            ],
            # ── Stage reorder
            STAGE_REORDER: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, stage_reorder_handler),
                back_mgmt, back,
            ],
            # ── Day field edit
            EDIT_DAY_NAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, edit_day_name_handler),
                back_mgmt, back,
            ],
            EDIT_DAY_FINAL: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, edit_day_final_handler),
                back_mgmt, back,
            ],
            EDIT_DAY_LOCKOUT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, edit_day_lockout_handler),
                back_mgmt, back,
            ],
            # ── Delete day
            DEL_DAY_SEL: [
                CallbackQueryHandler(del_day_sel, pattern=r"^delday_\d+$"),
                back,
            ],
            DEL_CONFIRM: [
                CallbackQueryHandler(del_confirm, pattern="^delday_confirm$"),
                back,
            ],
            # ── Reset
            RESET_CONFIRM: [
                CallbackQueryHandler(reset_confirm, pattern="^resetconfirm_yes$"),
                back,
            ],
            # ── Broadcast
            BROADCAST_MSG: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, broadcast_msg_handler),
                back,
            ],
            BROADCAST_CONFIRM: [
                CallbackQueryHandler(broadcast_confirm_handler, pattern="^bcastconfirm_yes$"),
                back,
            ],
            # ── Codes
            CODE_MANAGE: [
                CallbackQueryHandler(cb_codes_view,   pattern="^codes_view$"),
                CallbackQueryHandler(cb_codes_export, pattern="^codes_export$"),
                CallbackQueryHandler(cb_codes_create, pattern="^codes_create$"),
                CallbackQueryHandler(cb_codes,        pattern="^adm_codes$"),
                back,
            ],
            CODE_COUNT: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, codes_count_handler),
                CallbackQueryHandler(cb_codes, pattern="^adm_codes$"),
                back,
            ],
            CODE_POINTS: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, codes_points_handler),
                CallbackQueryHandler(cb_codes, pattern="^adm_codes$"),
                back,
            ],
            # ── Credit management
            CREDIT_SEARCH: [
                CallbackQueryHandler(credit_page_handler,        pattern=r"^credit_page_\d+$"),
                CallbackQueryHandler(credit_user_sel,            pattern=r"^credit_sel_\d+$"),
                CallbackQueryHandler(credit_search_mode_handler, pattern="^credit_search_mode$"),
                CallbackQueryHandler(credit_list_reset_handler,  pattern="^credit_list_reset$"),
                back,
            ],
            CREDIT_USER_SEL: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, credit_search_text_handler),
                CallbackQueryHandler(credit_cancel_search, pattern="^credit_cancel_search$"),
                back,
            ],
            CREDIT_USER_MENU: [
                CallbackQueryHandler(cb_credit_add,    pattern="^credit_add$"),
                CallbackQueryHandler(cb_credit_remove, pattern="^credit_remove$"),
                CallbackQueryHandler(cb_credit_reset,  pattern="^credit_reset$"),
                CallbackQueryHandler(cb_credit_log,    pattern="^credit_log$"),
                CallbackQueryHandler(credit_back,      pattern="^credit_back$"),
                back,
            ],
            CREDIT_ADD: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, credit_add_handler),
                CallbackQueryHandler(credit_back, pattern="^credit_back$"),
                back,
            ],
            CREDIT_REMOVE: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, credit_remove_handler),
                CallbackQueryHandler(credit_back, pattern="^credit_back$"),
                back,
            ],
            CREDIT_REMOVE_REASON: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, credit_remove_reason_handler),
                CallbackQueryHandler(credit_remove_skip, pattern="^credit_remove_skip$"),
                CallbackQueryHandler(credit_back,        pattern="^credit_back$"),
                back,
            ],
            CREDIT_RESET_CONFIRM: [
                CallbackQueryHandler(credit_reset_confirm, pattern="^credit_reset_yes$"),
                CallbackQueryHandler(credit_back,          pattern="^credit_back$"),
                back,
            ],
            # ── Participants list
            PART_LIST: [
                CallbackQueryHandler(part_page_handler,   pattern=r"^part_page_\d+$"),
                CallbackQueryHandler(part_detail_handler, pattern=r"^part_detail_\d+$"),
                back,
            ],
            PART_DETAIL: [
                CallbackQueryHandler(part_rename_start, pattern="^part_rename$"),
                CallbackQueryHandler(part_ban_start,    pattern="^part_ban$"),
                CallbackQueryHandler(part_unban,        pattern="^part_unban$"),
                CallbackQueryHandler(part_delete_start, pattern="^part_del$"),
                CallbackQueryHandler(part_back_to_list, pattern="^part_back_list$"),
                CallbackQueryHandler(part_detail_handler, pattern=r"^part_detail_\d+$"),
                back,
            ],
            PART_RENAME: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, part_rename_handler),
                CallbackQueryHandler(part_rename_cancel, pattern="^part_rename_cancel$"),
                back,
            ],
            PART_BAN_REASON: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, part_ban_reason_handler),
                CallbackQueryHandler(part_ban_skip,   pattern="^part_ban_skip$"),
                CallbackQueryHandler(part_ban_cancel, pattern="^part_ban_cancel$"),
                back,
            ],
            PART_DELETE_CONFIRM: [
                CallbackQueryHandler(part_delete_confirm, pattern="^part_del_yes$"),
                CallbackQueryHandler(part_delete_cancel,  pattern="^part_del_no$"),
                back,
            ],
            # ── Transaction log
            TRANS_LOG: [
                CallbackQueryHandler(tlog_user_start,   pattern="^tlog_user$"),
                CallbackQueryHandler(tlog_all_handler,  pattern="^tlog_all$"),
                CallbackQueryHandler(tlog_page_handler, pattern=r"^tlog_page_\d+$"),
                CallbackQueryHandler(tlog_upage_handler,pattern=r"^tlog_upage_\d+$"),
                CallbackQueryHandler(tlog_excel_handler,pattern="^tlog_excel$"),
                CallbackQueryHandler(tlog_pdf_handler,  pattern="^tlog_pdf$"),
                CallbackQueryHandler(tlog_delete_start, pattern="^tlog_delete$"),
                CallbackQueryHandler(cb_tlog_main,      pattern="^adm_tlog$"),
                back,
            ],
            TRANS_USER_SEARCH: [
                MessageHandler(filters.TEXT & ~filters.COMMAND, tlog_user_search_handler),
                CallbackQueryHandler(cb_tlog_main, pattern="^adm_tlog$"),
                back,
            ],
            TRANS_DELETE_CONFIRM: [
                CallbackQueryHandler(tlog_delete_confirm, pattern="^tlog_delete_yes$"),
                CallbackQueryHandler(cb_tlog_main,         pattern="^adm_tlog$"),
                back,
            ],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
        allow_reentry=True,
        per_message=False,
    )
